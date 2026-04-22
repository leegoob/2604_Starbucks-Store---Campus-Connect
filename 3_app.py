# -*- coding: utf-8 -*-
"""Streamlit: Kakao Map, 학교 필터, 복수 매장 탭, 학교별 근접매장 요약."""
from __future__ import annotations

import hmac
import html
import json
import math
import os
import re
import time
from collections import Counter, defaultdict
from io import BytesIO, StringIO
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = Path(__file__).resolve().parent
DATA = BASE / "processed_data.csv"
STORE_TAGS_FILE = BASE / "store_tags.csv"

APP_TITLE = "[인사기획팀] Starbucks Store - Campus Connect (2026 ver)"
PAGE_TITLE = APP_TITLE
STARBUCKS_GREEN = "#00704A"
# 표 셀 강조: 진한 초록(매장명·학교명) / 옅은 초록(순위별 학교명·매장 요약)
_TABLE_CELL_GREEN_DARK_BG = "#a3d6b4"
_TABLE_CELL_GREEN_DARK_FG = "#0d3d28"
_TABLE_CELL_GREEN_LIGHT_BG = "#e8f5ec"
_TABLE_CELL_GREEN_LIGHT_FG = "#1a5c38"
_SCHOOL_CENTRIC_NUMERIC_SORT_COLS = frozenset({"연관_매장_수"})
PRETENDARD_CSS = "https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css"
FONT_STACK = '"Pretendard Variable", Pretendard, -apple-system, BlinkMacSystemFont, system-ui, sans-serif'
FONT_STACK_HTML = "'Pretendard Variable', Pretendard, -apple-system, BlinkMacSystemFont, system-ui, sans-serif"
# 위젯·코드 느낌 보조 글꼴 (데이터 입력/콤보)
FONT_MONO = (
    "ui-monospace, 'SF Mono', 'Cascadia Code', 'Consolas', 'Malgun Gothic', monospace"
)

APP_SHELL_CSS = """
:root { --app-border: #c5cdd6; --app-surface: #ffffff; --app-canvas: #e9ecf0; --app-sidebar: #eceff3; }
section[data-testid="stSidebar"] {
  background: var(--app-sidebar) !important;
  border-right: 1px solid var(--app-border) !important;
}
[data-testid="stAppViewContainer"] { background: var(--app-canvas) !important; }
[data-testid="stHeader"] {
  background: rgba(248,249,251,0.92) !important;
  border-bottom: 1px solid var(--app-border) !important;
  backdrop-filter: blur(6px);
}
section.main [data-testid="stMainBlockContainer"] {
  padding-top: 1.1rem !important;
  padding-bottom: 2rem !important;
}
.stTextInput input, .stTextInput textarea,
div[data-baseweb="select"] > div, div[data-baseweb="input"] input {
  font-family: """ + FONT_MONO + r""" !important;
  font-size: 13px !important;
  border-radius: 4px !important;
  border-color: #adb5bd !important;
}
div[data-baseweb="select"] { background: var(--app-surface) !important; }
.app-section-title {
  font-size: 0.68rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.1em;
  color: #5c656f;
  margin: 0 0 0.35rem 0;
  font-family: """ + FONT_STACK + r""" !important;
}
.app-panel {
  border: 1px solid var(--app-border);
  border-radius: 6px;
  background: var(--app-surface);
  padding: 0.85rem 1rem 0.95rem 1rem;
  margin: 0 0 1rem 0;
  box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
}
.app-micro-hint {
  margin: 0.15rem 0 0.1rem 0;
  font-size: 0.72rem;
  color: #6b7280;
  font-family: """ + FONT_MONO + r""" !important;
}
[data-testid="stMetricValue"] { font-variant-numeric: tabular-nums; }
"""

# 매장·지도 탭: 처음엔 아무 매장도 고르지 않음 (CSV 첫 행으로 지도가 뜨는 것 방지)
STORE_PICK_NONE = -1
STORE_PICK_LABEL_NONE = "— 매장을 선택하세요 (선택 후 인근 학교·지도) —"

SCHOOL_CONTACT_COLUMNS: list[tuple[str, str]] = [
    ("campus_kind", "캠퍼스구분"),
    ("contact_office", "담당부서"),
    ("contact_name", "담당자명"),
    ("contact_phone", "전화번호"),
    ("contact_email", "이메일"),
]

STORE_META_INTERNAL = ("ops_team", "store_region")
STORE_TAG_COLUMNS = ("격오지", "채용난해", "중점관리", "TO미충원", "기타")
STORE_TAG_LEGACY_ALIASES = {
    "격오지 매장": "격오지",
    "채용난해매장": "채용난해",
    "중점관리매장": "중점관리",
    "TO미충원 매장": "TO미충원",
    "기타 매장": "기타",
}

KAKAO_REST_ADDRESS_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KAKAO_REST_KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"

# import_excel_to_csv.py 와 동일한 매장 엑셀 규칙
UPLOAD_STORE_RENAME: dict[str, str] = {
    "매장명": "name",
    "주소": "address",
}
UPLOAD_STORE_OPTIONAL_RENAME: dict[str, str] = {
    "운영팀": "ops_team",
    "권역": "store_region",
}

# 엑셀(.xlsx) 업로드 상한 — 서버 부하·실수로 인한 대용량 업로드 완화
_MAX_XLSX_UPLOAD_BYTES = 15 * 1024 * 1024


def _xlsx_upload_size_error(uploaded_file: object | None) -> str | None:
    """업로드 파일이 상한을 넘으면 사용자용 메시지, 아니면 None."""
    if uploaded_file is None:
        return None
    try:
        sz = int(getattr(uploaded_file, "size", 0) or 0)
    except (TypeError, ValueError):
        sz = 0
    if sz <= 0:
        try:
            sz = len(uploaded_file.getvalue())
        except Exception:
            return None
    if sz > _MAX_XLSX_UPLOAD_BYTES:
        mb = _MAX_XLSX_UPLOAD_BYTES // (1024 * 1024)
        return f"파일이 너무 큽니다({mb}MB 이하의 .xlsx만 업로드할 수 있습니다)."
    return None


def load_kakao_rest_key() -> str:
    k = os.environ.get("KAKAO_REST_API_KEY", "").strip()
    if k:
        return k
    try:
        return str(st.secrets.get("kakao_rest_api_key", "") or "").strip()
    except Exception:
        return ""


def _normalize_addr_app(s: str) -> str:
    s = str(s or "").strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    return re.sub(r"\s+", " ", s)


def _kakao_geocode_address(
    session: requests.Session,
    headers: dict[str, str],
    address: str,
) -> tuple[float | None, float | None]:
    addr = _normalize_addr_app(address)
    if not addr:
        return None, None
    for attempt in range(3):
        try:
            r = session.get(
                KAKAO_REST_ADDRESS_URL,
                headers=headers,
                params={"query": addr, "size": 1},
                timeout=20,
            )
            if r.status_code == 429:
                time.sleep(1.0 + attempt)
                continue
            r.raise_for_status()
            docs = (r.json() or {}).get("documents") or []
            if not docs:
                return None, None
            d0 = docs[0]
            return float(d0["y"]), float(d0["x"])
        except (requests.RequestException, KeyError, ValueError, TypeError):
            time.sleep(0.3 * (attempt + 1))
    return None, None


def _kakao_geocode_keyword(
    session: requests.Session,
    headers: dict[str, str],
    query: str,
) -> tuple[float | None, float | None]:
    q = _normalize_addr_app(query)[:100]
    if not q:
        return None, None
    for attempt in range(2):
        try:
            r = session.get(
                KAKAO_REST_KEYWORD_URL,
                headers=headers,
                params={"query": q, "size": 5},
                timeout=20,
            )
            if r.status_code == 429:
                time.sleep(1.0 + attempt)
                continue
            r.raise_for_status()
            docs = (r.json() or {}).get("documents") or []
            if not docs:
                return None, None
            d0 = docs[0]
            return float(d0["y"]), float(d0["x"])
        except (requests.RequestException, KeyError, ValueError, TypeError):
            time.sleep(0.3 * (attempt + 1))
    return None, None


def _resolve_uploaded_store_coords(
    session: requests.Session,
    headers: dict[str, str],
    name: str,
    address: str,
    pause: float,
) -> tuple[float | None, float | None]:
    lat, lon = _kakao_geocode_address(session, headers, address)
    if lat is not None and lon is not None:
        return lat, lon
    time.sleep(pause)
    lat, lon = _kakao_geocode_keyword(session, headers, name)
    if lat is not None and lon is not None:
        return lat, lon
    time.sleep(pause)
    full = _normalize_addr_app(address)
    short = full[:40] if full else ""
    if short and short != full:
        return _kakao_geocode_address(session, headers, short)
    return None, None


def _first_coord_column(columns: list, candidates: tuple[str, ...]) -> str | None:
    seen = {str(c).strip(): c for c in columns}
    seen_l = {str(c).strip().lower(): c for c in columns}
    for cand in candidates:
        if cand in seen:
            return str(seen[cand])
        cl = cand.lower()
        if cl in seen_l:
            return str(seen_l[cl])
    return None


def parse_uploaded_stores_excel(file_bytes: bytes) -> pd.DataFrame:
    """매장명·주소 필수. 위도/경도 열이 있으면 숫자로 채웁니다."""
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
    raw = raw.dropna(how="all")
    cols = list(raw.columns)
    for c in raw.columns:
        if raw[c].dtype == object:
            raw[c] = raw[c].astype(str).str.strip()
    lat_src = _first_coord_column(cols, ("위도", "latitude", "lat", "LAT", "y"))
    lng_src = _first_coord_column(cols, ("경도", "longitude", "lng", "lon", "LNG", "x"))
    lat_series = (
        pd.to_numeric(raw[lat_src], errors="coerce") if lat_src and lat_src in raw.columns else None
    )
    lng_series = (
        pd.to_numeric(raw[lng_src], errors="coerce") if lng_src and lng_src in raw.columns else None
    )
    df = raw.rename(columns={**UPLOAD_STORE_RENAME, **UPLOAD_STORE_OPTIONAL_RENAME})
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]
    for col in ("name", "address"):
        if col not in df.columns:
            raise ValueError(
                f"매장 엑셀에 필수 열이 없습니다: '{col}'. 헤더 예: 매장명, 주소. 현재: {list(df.columns)}"
            )
    opt_order = ["ops_team", "store_region"]
    present = [c for c in opt_order if c in df.columns]
    out = df[["name", "address"] + present].copy()
    out = out[out["name"].str.len() > 0]
    out = out[out["address"].str.len() > 0]
    if lat_series is not None:
        out["latitude"] = lat_series.reindex(out.index).to_numpy()
    else:
        out["latitude"] = np.nan
    if lng_series is not None:
        out["longitude"] = lng_series.reindex(out.index).to_numpy()
    else:
        out["longitude"] = np.nan
    out["latitude"] = pd.to_numeric(out["latitude"], errors="coerce")
    out["longitude"] = pd.to_numeric(out["longitude"], errors="coerce")
    for c in STORE_META_INTERNAL:
        if c not in out.columns:
            out[c] = ""
        else:
            out[c] = out[c].fillna("").astype(str).str.strip()
    return out.reset_index(drop=True)


def geocode_uploaded_stores(df: pd.DataFrame, rest_key: str, pause: float = 0.09) -> pd.DataFrame:
    out = df.copy()
    session = requests.Session()
    headers = {"Authorization": f"KakaoAK {rest_key}"}
    for idx in out.index:
        la = out.at[idx, "latitude"]
        lo = out.at[idx, "longitude"]
        if pd.notna(la) and pd.notna(lo):
            continue
        name = str(out.at[idx, "name"])
        addr = str(out.at[idx, "address"])
        nlat, nlng = _resolve_uploaded_store_coords(session, headers, name, addr, pause)
        if nlat is not None and nlng is not None:
            out.at[idx, "latitude"] = nlat
            out.at[idx, "longitude"] = nlng
    return out


def _ensure_store_meta(stores: pd.DataFrame) -> pd.DataFrame:
    s = stores.copy()
    for c in STORE_META_INTERNAL:
        if c not in s.columns:
            s[c] = ""
        else:
            s[c] = s[c].fillna("").astype(str).str.strip()
            s.loc[s[c].str.lower() == "nan", c] = ""
    return s


def _norm_tag_val(v: object) -> str:
    t = str(v or "").strip().upper()
    return "Y" if t in ("Y", "TRUE", "1", "예", "O") else "N"


def apply_store_tags(stores: pd.DataFrame) -> pd.DataFrame:
    out = stores.copy()
    for c in STORE_TAG_COLUMNS:
        if c not in out.columns:
            out[c] = "N"
        else:
            out[c] = out[c].map(_norm_tag_val)
    if not STORE_TAGS_FILE.is_file():
        return out
    try:
        tags = pd.read_csv(STORE_TAGS_FILE, encoding="utf-8-sig")
    except Exception:
        return out
    for old, new in STORE_TAG_LEGACY_ALIASES.items():
        if old in tags.columns and new not in tags.columns:
            tags[new] = tags[old]
    for k in ("name", "address"):
        if k not in tags.columns:
            return out
    base_cols = ["name", "address"] + list(STORE_TAG_COLUMNS)
    t = tags.reindex(columns=base_cols).copy()
    for c in STORE_TAG_COLUMNS:
        t[c] = t[c].map(_norm_tag_val)
    out = out.merge(t, on=["name", "address"], how="left", suffixes=("", "_tag"))
    for c in STORE_TAG_COLUMNS:
        tc = f"{c}_tag"
        if tc in out.columns:
            out[c] = out[tc].fillna(out[c]).map(_norm_tag_val)
            out = out.drop(columns=[tc])
    return out


def save_store_tags(stores: pd.DataFrame) -> None:
    cols = ["name", "address"] + list(STORE_TAG_COLUMNS)
    t = stores.reindex(columns=cols).copy()
    for c in STORE_TAG_COLUMNS:
        t[c] = t[c].map(_norm_tag_val)
    t.to_csv(STORE_TAGS_FILE, index=False, encoding="utf-8-sig")


def build_store_tags_template(stores: pd.DataFrame) -> pd.DataFrame:
    s = _ensure_store_meta(stores).copy().reset_index(drop=True)
    for c in STORE_TAG_COLUMNS:
        if c not in s.columns:
            s[c] = "N"
        s[c] = s[c].map(_norm_tag_val)
    out = s.rename(
        columns={
            "name": "매장명",
            "ops_team": "운영팀",
            "store_region": "권역",
        }
    )
    out["_team_sort"] = out["운영팀"].astype(str).map(_team_sort_key)
    out = out.sort_values(["_team_sort", "운영팀", "권역", "매장명"]).drop(columns=["_team_sort"]).reset_index(drop=True)
    cols = ["운영팀", "권역", "매장명"] + list(STORE_TAG_COLUMNS)
    return out[cols].copy()


def apply_store_tags_upload(file_bytes: bytes, stores: pd.DataFrame) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
    rename = {
        "매장": "매장명",
        "매장 이름": "매장명",
        "주소": "매장주소",
    }
    rename.update(STORE_TAG_LEGACY_ALIASES)
    df = raw.rename(columns=rename).copy()
    if "매장명" not in df.columns:
        raise ValueError("업로드 파일에 필수 열이 없습니다: 매장명")
    for c in STORE_TAG_COLUMNS:
        if c not in df.columns:
            df[c] = "N"
    tag_df = df[["매장명"] + list(STORE_TAG_COLUMNS)].copy()
    tag_df["매장명"] = tag_df["매장명"].fillna("").astype(str).str.strip()
    tag_df = tag_df[tag_df["매장명"] != ""]
    tag_df = tag_df.drop_duplicates(subset=["매장명"], keep="last")
    for c in STORE_TAG_COLUMNS:
        tag_df[c] = tag_df[c].map(_norm_tag_val)
    base = stores.copy().reset_index(drop=True)
    merged = base.merge(
        tag_df.rename(columns={"매장명": "name"}),
        on=["name"],
        how="left",
        suffixes=("", "_u"),
    )
    for c in STORE_TAG_COLUMNS:
        uc = f"{c}_u"
        merged[c] = merged[uc].fillna(merged.get(c, "N")).map(_norm_tag_val)
        merged = merged.drop(columns=[uc], errors="ignore")
    return merged


SCHOOL_FILTER_GROUPS: list[tuple[str, list[tuple[str, str]]]] = [
    (
        "고등학교",
        [
            ("hs_general", "일반고"),
            ("hs_autonomous", "자율고"),
            ("hs_special", "특목고"),
            ("hs_specialized", "특성화고"),
        ],
    ),
    (
        "대학교",
        [
            ("univ_4year", "4년제"),
            ("univ_junior", "전문대"),
            ("univ_cyber", "사이버대"),
            ("univ_other", "기타대(예: 기능대학 등)"),
        ],
    ),
]

SCHOOL_FILTER_DEF: list[tuple[str, str]] = [pair for _, pairs in SCHOOL_FILTER_GROUPS for pair in pairs]
SCHOOL_FILTER_LABEL_BY_KEY: dict[str, str] = {k: v for k, v in SCHOOL_FILTER_DEF}
HIDE_CONTACT_COLUMNS = ("담당부서", "담당자명", "전화번호", "이메일")


def school_filter_key(school_type: str) -> str:
    """학교구분 → 고등(일반·특목·특성화) / 대학(4년제·전문대·사이버·기타대) / 기타.

    공공·교육부 자료는 '일반고', '대학교' 등 짧은 표기인 경우가 많아 단축·정식 모두 처리합니다.
    """
    t = str(school_type or "").strip()
    if not t:
        return "other"
    # 정규화된 표준 라벨 우선 처리
    if t == "일반고":
        return "hs_general"
    if t == "특목고":
        return "hs_special"
    if t == "자율고":
        return "hs_autonomous"
    if t == "특성화고":
        return "hs_specialized"
    if t == "4년제":
        return "univ_4year"
    if t == "전문대":
        return "univ_junior"
    if t == "사이버대":
        return "univ_cyber"
    if t == "기타대":
        return "univ_other"
    if "특목고" in t:
        return "hs_special"
    if "특성화고" in t:
        return "hs_specialized"
    if "자율고" in t or "자율형" in t:
        return "hs_autonomous"
    if "일반고" in t:
        return "hs_general"
    if "고등학교" in t:
        if "특목" in t:
            return "hs_special"
        if "특성화" in t:
            return "hs_specialized"
        if "자율" in t:
            return "hs_autonomous"
        return "hs_general"
    if "외국인학교" in t and "고등" in t:
        if "특목" in t:
            return "hs_special"
        if "특성화" in t:
            return "hs_specialized"
        if "자율" in t:
            return "hs_autonomous"
        return "hs_general"
    if "방송통신" in t:
        return "univ_junior"
    if "사이버" in t or "디지털" in t:
        return "univ_cyber"
    if "전문대학" in t:
        return "univ_junior"
    if "기술대학" in t:
        return "univ_junior"
    if (
        "대학교" in t
        or "산업대학" in t
        or "교육대학" in t
        or "사범대학교" in t
    ):
        return "univ_4year"
    if t in ("대학", "일반대학"):
        return "univ_4year"
    if "대학원" in t:
        return "univ_4year"
    if "기능대학" in t or "각종학교" in t:
        return "univ_other"
    if "대학" in t:
        return "univ_other"
    return "other"


def normalize_school_type_value(v: object) -> str:
    """원본 school_type 값을 앱 전역 표준 라벨로 정규화."""
    key = school_filter_key(str(v or ""))
    if key == "univ_cyber":
        return "사이버대"
    if key == "univ_4year":
        return "4년제"
    if key == "univ_junior":
        return "전문대"
    if key == "hs_general":
        return "일반고"
    if key == "hs_special":
        return "특목고"
    if key == "hs_autonomous":
        return "자율고"
    if key == "hs_specialized":
        return "특성화고"
    if key == "univ_other":
        return "기타대"
    return str(v or "").strip()


def filter_schools_by_keys(schools: pd.DataFrame, selected_keys: set[str]) -> pd.DataFrame:
    if not selected_keys:
        return schools.copy()
    mask = schools["school_type"].map(school_filter_key).isin(selected_keys)
    return schools[mask].copy()


def school_key_counts(schools: pd.DataFrame) -> dict[str, int]:
    if schools.empty or "school_type" not in schools.columns:
        return {fk: 0 for fk, _ in SCHOOL_FILTER_DEF}
    vc = schools["school_type"].map(school_filter_key).value_counts()
    return {fk: int(vc.get(fk, 0)) for fk, _ in SCHOOL_FILTER_DEF}


def _school_centric_col_config(df: pd.DataFrame) -> dict[str, object]:
    """학교별 근접매장 요약 표: 연관_매장_수는 숫자 열로 두어 정렬이 1,2,… 순이 되게 함."""
    cfg = _table_col_config(df)
    if "연관_매장_수" in df.columns:
        cfg["연관_매장_수"] = st.column_config.NumberColumn(
            "연관_매장_수",
            format="%d",
            width=52,
        )
    return cfg


def simplify_school_name(v: object) -> str:
    n = str(v or "").strip()
    if not n:
        return ""
    repl = (
        ("여자고등학교", "여고"),
        ("고등학교", "고"),
        ("전문대학", "전문대"),
        ("대학교", "대"),
        ("여자중학교", "여중"),
        ("중학교", "중"),
        ("여자초등학교", "여초"),
        ("초등학교", "초"),
    )
    out = n
    for a, b in repl:
        out = out.replace(a, b)
    return out


def _school_row_key(name: object, address: object) -> str:
    """근접·지도에서 학교 행을 구분 (동명이면서 주소가 다른 경우 포함). 표시용이 아님."""
    return f"{str(name or '').strip()}\x1f{str(address or '').strip()}"


def _near_school_table_label(row: pd.Series) -> str:
    """인근 학교 표/지도 선택용 라벨: 학교명 · 주소."""
    nm = str(row.get("학교명", "")).strip()
    addr = str(row.get("주소", "")).strip()
    if len(addr) > 40:
        addr = addr[:38] + "…"
    return f"{nm} · {addr}" if addr else nm


def school_type_display_label(v: object) -> str:
    key = school_filter_key(str(v or ""))
    return SCHOOL_FILTER_LABEL_BY_KEY.get(key, "기타")


def _sanitize_table(df: pd.DataFrame, *, hide_store_address: bool = False) -> pd.DataFrame:
    out = df.copy()
    drop_cols = [c for c in HIDE_CONTACT_COLUMNS if c in out.columns]
    if hide_store_address and "매장주소" in out.columns:
        drop_cols.append("매장주소")
    if drop_cols:
        out = out.drop(columns=drop_cols, errors="ignore")
    for c in out.columns:
        cn = str(c)
        if "(km)" in cn or cn.endswith("_km"):
            num = pd.to_numeric(out[c], errors="coerce")
            if num.notna().any():
                out[c] = num.round(1)
    return out


def _format_table_cells_for_display(
    view: pd.DataFrame,
    *,
    keep_numeric_for_sort: frozenset[str] | None = None,
) -> None:
    """render_table과 동일한 표시용 문자열 변환(in-place).

    keep_numeric_for_sort: 이 열들은 숫자 dtype을 유지해 그리드 정렬이 숫자 기준으로 동작하게 함.
    """
    keep = keep_numeric_for_sort or frozenset()
    for c in view.columns:
        if c == "No":
            view[c] = pd.to_numeric(view[c], errors="coerce").fillna(0).astype(int)
            continue
        if str(c) in keep:
            view[c] = pd.to_numeric(view[c], errors="coerce").fillna(0).astype(int)
            continue
        if pd.api.types.is_numeric_dtype(view[c]):
            if "(km)" in str(c) or str(c).endswith("_km"):
                view[c] = pd.to_numeric(view[c], errors="coerce").map(lambda x: "" if pd.isna(x) else f"{x:.1f}")
            elif "추정" in str(c) and "분" in str(c):
                view[c] = pd.to_numeric(view[c], errors="coerce").map(lambda x: "" if pd.isna(x) else f"{x:.1f}")
            else:
                view[c] = pd.to_numeric(view[c], errors="coerce").map(lambda x: "" if pd.isna(x) else f"{x:g}")
        else:
            view[c] = view[c].astype(str).replace({"nan": "", "None": ""})


def render_table(df: pd.DataFrame, *, height: int | None = None, use_container_width: bool = False) -> None:
    view = _sanitize_table(df, hide_store_address=True)
    if "No" not in view.columns:
        view = view.reset_index(drop=True)
        view.insert(0, "No", np.arange(1, len(view) + 1))
    _format_table_cells_for_display(view)
    kwargs: dict[str, object] = {
        "use_container_width": use_container_width,
        "hide_index": True,
        "column_config": _table_col_config(view),
    }
    if height is not None:
        kwargs["height"] = int(height)
    st.dataframe(view, **kwargs)


_WID_RANK_SCHOOL_NAME_RE = re.compile(r"^[1-9]\d*위_학교명$")
# 매장·지도 인근 표 등: N위_열 폭
_RANK_COL_PREFIX_RE = re.compile(r"^[1-9]\d*위_")


def _table_col_config(df: pd.DataFrame) -> dict[str, object]:
    """표 컬럼 폭. 순번·거리·순위 열은 짧게, 이름/주소는 내용에 맞추되 과도한 확장 억제."""
    cfg: dict[str, object] = {}
    if df.empty:
        return cfg
    sample = df.head(300)
    for col in df.columns:
        c_name = str(col)
        if c_name == "No":
            cfg[c_name] = st.column_config.NumberColumn(
                c_name,
                format="%d",
                width=30,
            )
            continue
        if "(km)" in c_name or "직선거리" in c_name or c_name.endswith("_km"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=78)
            continue
        if "추정" in c_name and "분" in c_name:
            cfg[c_name] = st.column_config.TextColumn(c_name, width=92)
            continue
        if _RANK_COL_PREFIX_RE.match(c_name):
            if "학교명" in c_name:
                cfg[c_name] = st.column_config.TextColumn(c_name, width=118)
            elif "주소" in c_name:
                cfg[c_name] = st.column_config.TextColumn(c_name, width=176)
            else:
                cfg[c_name] = st.column_config.TextColumn(c_name, width=92)
            continue
        if c_name in ("학교명", "매장명"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=136)
            continue
        if "주소" in c_name or c_name in ("address", "매장주소", "학교주소"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=188)
            continue
        if c_name in ("학교구분", "학교유형", "캠퍼스구분", "캠퍼스", "특성구분"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=84)
            continue
        if c_name in ("운영팀", "권역", "지역"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=92)
            continue
        if c_name in ("담당부서", "담당자명", "전화번호", "이메일"):
            cfg[c_name] = st.column_config.TextColumn(c_name, width=96)
            continue
        s = sample[col]
        max_len = max(
            len(c_name),
            int(s.astype(str).map(len).max()) if len(s) else 0,
        )
        if max_len <= 5:
            w_px = 64
        elif max_len <= 12:
            w_px = 88
        elif max_len <= 22:
            w_px = 120
        elif max_len <= 36:
            w_px = 152
        else:
            w_px = min(220, 100 + int(max_len * 3.2))
        cfg[c_name] = st.column_config.TextColumn(c_name, width=int(w_px))
    return cfg


def render_wide_store_school_table(
    df: pd.DataFrame,
    *,
    height: int | None = 480,
    use_container_width: bool = True,
) -> None:
    """매장별 근접학교 요약: 매장명=진한 초록, 순위별 학교명 열=옅은 초록."""
    view = _sanitize_table(df, hide_store_address=True)
    if "No" not in view.columns:
        view = view.reset_index(drop=True)
        view.insert(0, "No", np.arange(1, len(view) + 1))
    _format_table_cells_for_display(view)
    school_cols = [c for c in view.columns if _WID_RANK_SCHOOL_NAME_RE.match(str(c))]
    dark_store = (
        f"background-color: {_TABLE_CELL_GREEN_DARK_BG}; color: {_TABLE_CELL_GREEN_DARK_FG}; font-weight: 600"
    )
    light_school = (
        f"background-color: {_TABLE_CELL_GREEN_LIGHT_BG}; color: {_TABLE_CELL_GREEN_LIGHT_FG}"
    )
    styled = view.style
    if "매장명" in view.columns:
        styled = styled.map(lambda _: dark_store, subset=["매장명"])
    if school_cols:
        styled = styled.map(lambda _: light_school, subset=school_cols)
    kwargs: dict[str, object] = {
        "use_container_width": use_container_width,
        "hide_index": True,
        "column_config": _table_col_config(view),
    }
    if height is not None:
        kwargs["height"] = int(height)
    st.dataframe(styled, **kwargs)


def render_school_centric_table(
    df: pd.DataFrame,
    *,
    height: int | None = 360,
    use_container_width: bool = True,
) -> None:
    """학교별 근접매장 요약: 학교명=진한 초록, 매장_요약=옅은 초록, 연관_매장_수는 숫자 정렬."""
    view = _sanitize_table(df, hide_store_address=True)
    if "No" not in view.columns:
        view = view.reset_index(drop=True)
        view.insert(0, "No", np.arange(1, len(view) + 1))
    _format_table_cells_for_display(view, keep_numeric_for_sort=_SCHOOL_CENTRIC_NUMERIC_SORT_COLS)
    dark = (
        f"background-color: {_TABLE_CELL_GREEN_DARK_BG}; color: {_TABLE_CELL_GREEN_DARK_FG}; font-weight: 600"
    )
    light = (
        f"background-color: {_TABLE_CELL_GREEN_LIGHT_BG}; color: {_TABLE_CELL_GREEN_LIGHT_FG}"
    )
    styled = view.style
    if "학교명" in view.columns:
        styled = styled.map(lambda _: dark, subset=["학교명"])
    if "매장_요약" in view.columns:
        styled = styled.map(lambda _: light, subset=["매장_요약"])
    kwargs: dict[str, object] = {
        "use_container_width": use_container_width,
        "hide_index": True,
        "column_config": _school_centric_col_config(view),
    }
    if height is not None:
        kwargs["height"] = int(height)
    st.dataframe(styled, **kwargs)


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(max(0.0, 1 - a)))
    return r * c


# 직선거리만 사용한 참고용 추정 소요(분). 실제 길찾기 API 아님(도심 유효속도·우회 가정).
_TRAVEL_ROAD_FACTOR = 1.3
_TRAVEL_CAR_KMH = 28.0
_TRAVEL_TRANSIT_KMH = 11.0

TRAVEL_TIME_HELP = (
    "자차·대중교통 「추정(분)」은 직선거리에 도심권 평균 유효속도·경로 왜곡을 반영한 "
    "추정치입니다. 실제 길찾기·혼잡·배차와 다를 수 있습니다."
)


def est_travel_minutes_from_straight_km(straight_km: float) -> tuple[float, float]:
    """직선거리(km) → (자차 분, 대중교통 분) 참고 추정."""
    try:
        d = float(straight_km)
    except (TypeError, ValueError):
        return (float("nan"), float("nan"))
    if d <= 0 or math.isnan(d):
        return (float("nan"), float("nan"))
    eff_km = max(d * _TRAVEL_ROAD_FACTOR, 0.05)
    car_m = (eff_km / _TRAVEL_CAR_KMH) * 60.0
    transit_m = (eff_km / _TRAVEL_TRANSIT_KMH) * 60.0
    return (round(car_m, 1), round(transit_m, 1))


def _haversine_matrix(
    slat: np.ndarray,
    slon: np.ndarray,
    lat: np.ndarray,
    lon: np.ndarray,
) -> np.ndarray:
    """shape (S,) × (K,) → 거리 행렬 (S, K) km. top5_for_store와 동일 공식."""
    slat_r = np.radians(slat)[:, None]
    slon_r = np.radians(slon)[:, None]
    lat_r = np.radians(lat)[None, :]
    lon_r = np.radians(lon)[None, :]
    dlat = lat_r - slat_r
    dlon = lon_r - slon_r
    sin_dlat = np.sin(dlat / 2.0)
    sin_dlon = np.sin(dlon / 2.0)
    a = sin_dlat * sin_dlat + np.cos(slat_r) * np.cos(lat_r) * sin_dlon * sin_dlon
    c = 2.0 * np.arctan2(np.sqrt(a), np.sqrt(np.maximum(0.0, 1.0 - a)))
    return 6371.0 * c


def _batch_topn_school_indices(
    stores: pd.DataFrame,
    schools: pd.DataFrame,
    n: int,
) -> tuple[np.ndarray, np.ndarray, pd.DataFrame]:
    """매장 행 순서와 동일하게, 각 행마다 가까운 학교 인덱스(n개 이하)와 직선거리(km).
    학교 서브프레임은 top5_for_store와 같이 위·경도 유효 행만 사용."""
    n_req = max(1, min(int(n), 50))
    sub = schools.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    s_count = len(stores)
    if sub.empty or s_count == 0:
        return (
            np.full((s_count, n_req), -1, dtype=np.int32),
            np.full((s_count, n_req), np.nan, dtype=np.float64),
            sub,
        )

    lat_s = sub["latitude"].astype(float).to_numpy()
    lon_s = sub["longitude"].astype(float).to_numpy()
    k = len(sub)
    n_take = min(n_req, k)

    slat_all = stores["latitude"].to_numpy(dtype=float, copy=False)
    slon_all = stores["longitude"].to_numpy(dtype=float, copy=False)
    valid_mask = ~(np.isnan(slat_all) | np.isnan(slon_all))
    valid_rows = np.flatnonzero(valid_mask)

    top_idx = np.full((s_count, n_take), -1, dtype=np.int32)
    top_dist = np.full((s_count, n_take), np.nan, dtype=np.float64)

    # 피크 메모리 ~ float64(S*K): 청크로 나눔 (대략 90MB 상한)
    max_cells = 9_000_000
    chunk = max(1, min(s_count, max(1, max_cells // max(k, 1))))

    for c0 in range(0, len(valid_rows), chunk):
        vr = valid_rows[c0 : c0 + chunk]
        slat = slat_all[vr]
        slon = slon_all[vr]
        dist = _haversine_matrix(slat, slon, lat_s, lon_s)
        if n_take == k:
            idx = np.argsort(dist, axis=1)
        else:
            part = np.argpartition(dist, n_take - 1, axis=1)[:, :n_take]
            row_ids = np.arange(len(vr), dtype=np.intp)[:, None]
            dpart = dist[row_ids, part]
            order = np.argsort(dpart, axis=1)
            idx = part[row_ids, order].astype(np.int32, copy=False)
        top_idx[vr] = idx
        top_dist[vr] = dist[np.arange(len(vr), dtype=np.intp)[:, None], idx]

    if n_take < n_req:
        top_idx = np.hstack(
            [top_idx, np.full((s_count, n_req - n_take), -1, dtype=np.int32)]
        )
        top_dist = np.hstack(
            [top_dist, np.full((s_count, n_req - n_take), np.nan, dtype=np.float64)]
        )
    return top_idx, top_dist, sub


@st.cache_data(show_spinner="매장–학교 근접 계산 중…")
def _cached_batch_neighbors(
    stores_view: pd.DataFrame,
    schools_use: pd.DataFrame,
    n_near: int,
) -> tuple[np.ndarray, np.ndarray, pd.DataFrame]:
    return _batch_topn_school_indices(
        stores_view.reset_index(drop=True),
        schools_use,
        n_near,
    )


@st.cache_data
def load_data(path: str, mtime_ns: int) -> pd.DataFrame:
    """mtime_ns로 파일이 바뀌면 캐시를 새로 읽습니다."""
    _ = mtime_ns
    tried: list[str] = []
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            tried.append(enc)
            continue
        except pd.errors.ParserError:
            # CSV 내 일부 깨진 줄(열 개수 불일치)은 건너뛰고 최대한 로드
            try:
                return pd.read_csv(path, encoding=enc, engine="python", on_bad_lines="skip")
            except Exception:
                tried.append(f"{enc}(parser)")
                continue
    # 마지막 안전장치: utf-8 기준으로 깨진 바이트를 치환해 최대한 로드
    try:
        txt = Path(path).read_text(encoding="utf-8-sig", errors="replace")
        return pd.read_csv(StringIO(txt), engine="python", on_bad_lines="skip")
    except Exception as e:
        raise ValueError(f"CSV 로딩 실패(인코딩/구분자/행손상). 시도: {', '.join(tried)} / {e}") from e


def top5_for_store(slat: float, slon: float, schools: pd.DataFrame, n: int = 5) -> pd.DataFrame:
    """거리 계산을 벡터화해 스크립트 재실행(흰 끊김) 시간을 줄입니다."""
    n = max(1, min(int(n), 50))
    sub = schools.dropna(subset=["latitude", "longitude"])
    _geo_cols = ["_geo_lat", "_geo_lng"]
    if sub.empty:
        return pd.DataFrame(
            columns=[
                "학교명",
                "학교구분",
                "직선거리(km)",
                "자차 추정(분)",
                "대중교통 추정(분)",
                "주소",
                "_school_key",
            ]
            + _geo_cols
            + [kr for _, kr in SCHOOL_CONTACT_COLUMNS]
        )
    lat = sub["latitude"].astype(float).to_numpy()
    lon = sub["longitude"].astype(float).to_numpy()
    slat_r = math.radians(slat)
    slon_r = math.radians(slon)
    lat_r = np.radians(lat)
    lon_r = np.radians(lon)
    dlat = lat_r - slat_r
    dlon = lon_r - slon_r
    sin_dlat = np.sin(dlat / 2.0)
    sin_dlon = np.sin(dlon / 2.0)
    a = sin_dlat * sin_dlat + np.cos(slat_r) * np.cos(lat_r) * sin_dlon * sin_dlon
    c = 2.0 * np.arctan2(np.sqrt(a), np.sqrt(np.maximum(0.0, 1.0 - a)))
    dist_km = 6371.0 * c
    sub = sub.assign(_dkm=dist_km)
    sub = sub.nsmallest(n, "_dkm")
    rows = []
    for _, s in sub.iterrows():
        d = float(s["_dkm"])
        _car_m, _tr_m = est_travel_minutes_from_straight_km(d)
        row = {
            "학교명": simplify_school_name(s.get("name", "")),
            "학교구분": school_type_display_label(s.get("school_type", "")),
            "직선거리(km)": round(d, 3),
            "자차 추정(분)": _car_m,
            "대중교통 추정(분)": _tr_m,
            "주소": s["address"],
            "_school_key": _school_row_key(s.get("name", ""), s.get("address", "")),
            "_geo_lat": float(s["latitude"]),
            "_geo_lng": float(s["longitude"]),
        }
        for internal, kr in SCHOOL_CONTACT_COLUMNS:
            if internal in schools.columns:
                v = s[internal]
                row[kr] = str(v).strip() if pd.notna(v) else ""
            else:
                row[kr] = ""
        rows.append(row)
    return pd.DataFrame(rows).reset_index(drop=True)


def topn_nearby_stores(
    slat: float,
    slon: float,
    stores: pd.DataFrame,
    *,
    n: int = 5,
    exclude_name: str = "",
    exclude_address: str = "",
) -> pd.DataFrame:
    """선택 매장 기준 가장 가까운 다른 매장 n개."""
    n = max(1, min(int(n), 50))
    sub = stores.dropna(subset=["latitude", "longitude"]).copy()
    if sub.empty:
        return pd.DataFrame(
            columns=[
                "매장명",
                "직선거리(km)",
                "자차 추정(분)",
                "대중교통 추정(분)",
                "매장주소",
                "권역",
                "_geo_lat",
                "_geo_lng",
            ]
        )
    ex_name = str(exclude_name or "").strip()
    ex_addr = str(exclude_address or "").strip()
    if ex_name and ex_addr:
        sub = sub[~((sub["name"].astype(str) == ex_name) & (sub["address"].astype(str) == ex_addr))]
    elif ex_name:
        sub = sub[sub["name"].astype(str) != ex_name]
    if sub.empty:
        return pd.DataFrame(
            columns=[
                "매장명",
                "직선거리(km)",
                "자차 추정(분)",
                "대중교통 추정(분)",
                "매장주소",
                "권역",
                "_geo_lat",
                "_geo_lng",
            ]
        )

    lat = sub["latitude"].astype(float).to_numpy()
    lon = sub["longitude"].astype(float).to_numpy()
    slat_r = math.radians(slat)
    slon_r = math.radians(slon)
    lat_r = np.radians(lat)
    lon_r = np.radians(lon)
    dlat = lat_r - slat_r
    dlon = lon_r - slon_r
    sin_dlat = np.sin(dlat / 2.0)
    sin_dlon = np.sin(dlon / 2.0)
    a = sin_dlat * sin_dlat + np.cos(slat_r) * np.cos(lat_r) * sin_dlon * sin_dlon
    c = 2.0 * np.arctan2(np.sqrt(a), np.sqrt(np.maximum(0.0, 1.0 - a)))
    dist_km = 6371.0 * c
    sub = sub.assign(_dkm=dist_km).nsmallest(n, "_dkm")
    rows: list[dict[str, object]] = []
    for _, r in sub.iterrows():
        d_km = float(r["_dkm"])
        _c_m, _t_m = est_travel_minutes_from_straight_km(d_km)
        rows.append(
            {
                "매장명": str(r.get("name", "")),
                "직선거리(km)": round(d_km, 3),
                "자차 추정(분)": _c_m,
                "대중교통 추정(분)": _t_m,
                "매장주소": str(r.get("address", "")),
                "권역": str(r.get("store_region", "")),
                "_geo_lat": float(r["latitude"]),
                "_geo_lng": float(r["longitude"]),
            }
        )
    return pd.DataFrame(rows).reset_index(drop=True)


def _store_search_match_priorities(df: pd.DataFrame, ql: str) -> np.ndarray:
    """행마다 0=매장명 1=주소 2=운영팀 3=권역 4=pick_label, 99=불일치. ql은 소문자."""
    n = len(df)
    if not ql:
        return np.full(n, 99, dtype=np.int32)
    addr = df["address"].astype(str).str.lower()
    name = df["name"].astype(str).str.lower()
    reg = (
        df["store_region"].astype(str).str.lower()
        if "store_region" in df.columns
        else pd.Series("", index=df.index, dtype=str)
    )
    team = (
        df["ops_team"].astype(str).str.lower()
        if "ops_team" in df.columns
        else pd.Series("", index=df.index, dtype=str)
    )
    lab = df["pick_label"].astype(str).str.lower()
    m_addr = addr.str.contains(ql, regex=False, na=False)
    m_reg = reg.str.contains(ql, regex=False, na=False)
    m_team = team.str.contains(ql, regex=False, na=False)
    m_name = name.str.contains(ql, regex=False, na=False)
    m_lab = lab.str.contains(ql, regex=False, na=False)
    p = np.full(n, 99, dtype=np.int32)
    for tier, mask in enumerate([m_name, m_addr, m_team, m_reg, m_lab]):
        m = mask.fillna(False).to_numpy(dtype=bool)
        p[m] = np.minimum(p[m], tier)
    return p


def suggest_store_pick_labels(df: pd.DataFrame, query: str, *, limit: int = 28) -> list[str]:
    """검색어 부분 일치. 우선순위: 매장명→주소→운영팀→권역(difflib 없음)."""
    q = (query or "").strip()
    if not q:
        return []
    p = _store_search_match_priorities(df, q.lower())
    if (p >= 99).all():
        return []
    sub = df.assign(_prio=p).loc[p < 99].sort_values(["_prio", "pick_label"])
    return sub["pick_label"].drop_duplicates().head(limit).tolist()


def build_store_pick_frame(stores: pd.DataFrame) -> pd.DataFrame:
    sp = _ensure_store_meta(stores).reset_index(drop=True)
    labels: list[str] = []
    rec_labels: list[str] = []
    counts: Counter[str] = Counter()
    rec_counts: Counter[str] = Counter()
    for _, row in sp.iterrows():
        name_addr = f"{str(row['name']).strip()} — {str(row.get('address', '')).strip()}"
        counts[name_addr] += 1
        dup = f" (#{counts[name_addr]})" if counts[name_addr] > 1 else ""
        ot = str(row.get("ops_team", "") or "").strip()
        rg = str(row.get("store_region", "") or "").strip()
        org = " · ".join(x for x in (ot, rg) if x)
        suffix = f" · {org}" if org else ""
        labels.append(f"{name_addr}{suffix}{dup}")
        nm = str(row["name"]).strip()
        rbase = f"{nm} · {ot}" if ot else nm
        rec_counts[rbase] += 1
        rdup = f" (#{rec_counts[rbase]})" if rec_counts[rbase] > 1 else ""
        rec_labels.append(f"{rbase}{rdup}")
    sp["pick_label"] = labels
    sp["recommend_label"] = rec_labels
    return sp


def build_all_stores_wide(stores: pd.DataFrame, schools: pd.DataFrame, n: int = 5) -> pd.DataFrame:
    n = max(1, min(int(n), 50))
    st_work = stores.reset_index(drop=True)
    top_idx, top_dist, sub = _cached_batch_neighbors(st_work, schools, n)
    out_rows: list[dict[str, object]] = []
    for i, strow in st_work.iterrows():
        active_tags = [
            tc for tc in STORE_TAG_COLUMNS
            if _norm_tag_val(strow.get(tc, "N")) == "Y"
        ]
        row: dict[str, object] = {
            "특성구분": " · ".join(active_tags) if active_tags else "",
            "매장명": strow["name"],
            "매장주소": strow["address"],
            "운영팀": str(strow.get("ops_team", "") or "").strip(),
            "권역": str(strow.get("store_region", "") or "").strip(),
        }
        for rank in range(n):
            prefix = f"{rank + 1}위"
            ji = int(top_idx[i, rank])
            dkm = top_dist[i, rank]
            if ji < 0 or np.isnan(dkm):
                row[f"{prefix}_학교명"] = ""
                row[f"{prefix}_구분"] = ""
                row[f"{prefix}_캠퍼스"] = ""
                row[f"{prefix}_직선거리(km)"] = ""
                row[f"{prefix}_주소"] = ""
            else:
                s = sub.iloc[ji]
                ck = s.get("campus_kind")
                campus = str(ck).strip() if ck is not None and pd.notna(ck) else ""
                row[f"{prefix}_학교명"] = simplify_school_name(s.get("name", ""))
                row[f"{prefix}_구분"] = school_type_display_label(s.get("school_type", ""))
                row[f"{prefix}_캠퍼스"] = campus
                row[f"{prefix}_직선거리(km)"] = round(float(dkm), 3)
                row[f"{prefix}_주소"] = s["address"]
        out_rows.append(row)
    return pd.DataFrame(out_rows)


def build_school_centric_table(stores: pd.DataFrame, schools: pd.DataFrame, n_near: int) -> pd.DataFrame:
    """학교별로 인근 상위 n_near 안에 포함되는 매장 목록.

    근접 계산의 학교 행(`sub`) 인덱스로 집계한다. 동일 학교명이 여러 행(다른 주소·캠퍼스)이어도 병합되지 않는다.
    """
    st_work = stores.reset_index(drop=True)
    top_idx, top_dist, sub = _cached_batch_neighbors(st_work, schools, n_near)
    inv: dict[int, list[dict[str, object]]] = defaultdict(list)
    for i, strow in st_work.iterrows():
        if pd.isna(strow.get("latitude")) or pd.isna(strow.get("longitude")):
            continue
        for j in range(top_idx.shape[1]):
            ji = int(top_idx[i, j])
            dkm = top_dist[i, j]
            if ji < 0 or np.isnan(dkm):
                continue
            inv[ji].append(
                {
                    "매장명": str(strow["name"]),
                    "직선거리(km)": round(float(dkm), 3),
                    "운영팀": str(strow.get("ops_team", "") or ""),
                    "권역": str(strow.get("store_region", "") or ""),
                }
            )
    rows_out: list[dict[str, object]] = []
    for ji, sch in sub.iterrows():
        if ji not in inv:
            continue
        lst = inv[int(ji)]
        lst.sort(key=lambda x: float(x["직선거리(km)"]))
        unique_stores = []
        seen: set[str] = set()
        for it in lst:
            k = str(it["매장명"])
            if k not in seen:
                seen.add(k)
                unique_stores.append(it)
        sn = str(sch["name"]).strip()
        snip = " · ".join([str(x["매장명"]) for x in unique_stores[:30]])
        if len(unique_stores) > 30:
            snip += f" … 외 {len(unique_stores) - 30}곳"
        row: dict[str, object] = {
            "학교명": simplify_school_name(sn),
            "학교구분": school_type_display_label(sch.get("school_type", "")),
            "캠퍼스구분": str(sch.get("campus_kind", "") or "").strip(),
            "주소": sch.get("address", ""),
            "연관_매장_수": len(unique_stores),
            "최대직선거리(km)": round(float(max([float(x["직선거리(km)"]) for x in unique_stores])), 3),
            "매장_요약": snip,
        }
        for internal, kr in SCHOOL_CONTACT_COLUMNS:
            if internal in sch.index:
                v = sch[internal]
                row[kr] = str(v).strip() if pd.notna(v) else ""
            else:
                row[kr] = ""
        rows_out.append(row)
    out = pd.DataFrame(rows_out)
    if out.empty:
        return out
    return out.sort_values(["연관_매장_수", "학교명", "주소"], ascending=[False, True, True]).reset_index(drop=True)


# 템플릿 기본: 매장명만. 구 양식(권역·우선순위 포함)도 parse에서 호환.
CAMPAIGN_TEMPLATE_COLUMNS = ["매장명"]
CAMPAIGN_EXEC_COLUMNS = [
    "권역",
    "우선순위",
    "매장명",
    "학교명",
    "학교유형",
    "학교주소",
    "직선거리(km)",
    "담당부서",
    "담당자명",
    "전화번호",
    "이메일",
    "연락상태",
    "메모",
]


def _team_key(s: str) -> str:
    t = str(s or "").strip()
    if "(" in t:
        return t.split("(", 1)[0].strip()
    return t


def _team_sort_key(s: str) -> tuple[int, int, str]:
    t = str(s or "").strip()
    m = re.search(r"운영\s*(\d+)", _team_key(t))
    if m:
        return (0, int(m.group(1)), t)
    return (1, 10**9, t)


def _region_key(s: str) -> str:
    return str(s or "").strip().lower()


def build_campaign_template(blank_rows: int = 40) -> pd.DataFrame:
    """헤더 «매장명» + 빈 행. 권역·우선순위는 데이터/파일순으로 자동."""
    n = max(8, min(int(blank_rows), 500))
    return pd.DataFrame([{"매장명": ""} for _ in range(n)], columns=CAMPAIGN_TEMPLATE_COLUMNS)


def parse_campaign_submission_xlsx(file_bytes: bytes) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl").dropna(how="all")
    rename_map = {
        "권역명": "권역",
        "지역": "권역",
        "운영구역": "권역",
        "운영팀": "권역",  # 과거 양식 호환
        "운영팀명": "권역",  # 과거 양식 호환
        "순위": "우선순위",
        "우선 순위": "우선순위",
        "매장": "매장명",
        "매장 이름": "매장명",
    }
    df = raw.rename(columns=rename_map).copy()
    if "매장명" not in df.columns:
        raise ValueError("업로드 파일에 «매장명» 열이 필요합니다.")
    if "권역" not in df.columns:
        df["권역"] = ""
    if "우선순위" not in df.columns:
        df["우선순위"] = pd.NA
    out = df[["권역", "우선순위", "매장명"]].copy()
    for c in ("권역", "매장명"):
        out[c] = out[c].fillna("").astype(str).str.strip()
    out["우선순위"] = pd.to_numeric(out["우선순위"], errors="coerce").astype("Int64")
    # 매장명이 있는 행만 (권역은 비어 있어도 됨 → 마스터 권역 사용)
    out = out[out["매장명"].str.len() > 0]
    return out.reset_index(drop=True)


def resolve_campaign_stores(
    plan_df: pd.DataFrame,
    stores_master: pd.DataFrame,
    expected_per_region: int = 3,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    """매장명만 있어도 됨. 권역은 마스터(store_region), 우선순위는 파일 순서(권역별 1..n, n은 상한)."""
    warnings: list[str] = []
    errors: list[str] = []
    if plan_df.empty:
        return pd.DataFrame(), ["업로드 파일에 입력된 행이 없습니다."], []

    work = plan_df.copy()
    n = max(1, min(int(expected_per_region), 10))
    master = _ensure_store_meta(stores_master).copy().reset_index(drop=True)
    master["name_norm"] = master["name"].astype(str).str.strip()
    master_regions = [
        str(x).strip() for x in master.get("store_region", pd.Series(dtype=str)).tolist() if str(x).strip()
    ]
    canonical_by_key: dict[str, str] = {}
    for rg in master_regions:
        canonical_by_key.setdefault(_region_key(rg), rg)
    work["name_norm"] = work["매장명"].astype(str).str.strip()
    work["권역"] = work["권역"].fillna("").astype(str).str.strip()

    pending: list[tuple[int, pd.Series, str, object]] = []
    for orig_i, (_, r) in enumerate(work.iterrows()):
        nm = str(r["name_norm"])
        if not nm:
            continue
        cand = master.loc[master["name_norm"] == nm]
        if cand.empty:
            errors.append(f"매장 불일치: {r['매장명']}")
            continue
        pick = cand.iloc[0].copy()
        raw_region = str(r["권역"]).strip()
        real_region = str(pick.get("store_region", "") or "").strip()
        ex_rank = r.get("우선순위", pd.NA)
        if raw_region:
            req_region = canonical_by_key.get(_region_key(raw_region), raw_region)
            if req_region and real_region and _region_key(req_region) != _region_key(real_region):
                errors.append(
                    f"권역 불일치: 입력[{req_region}] / 매장[{pick.get('name', '')}] 실제권역[{real_region}]"
                )
                continue
            camp_region = req_region or real_region
        else:
            if not real_region:
                errors.append(f"데이터에 권역 없음: {pick.get('name', '')}")
                continue
            camp_region = canonical_by_key.get(_region_key(real_region), real_region)
        pending.append((orig_i, pick, camp_region, ex_rank))

    if not pending:
        return pd.DataFrame(), warnings, errors

    by_reg: dict[str, list[tuple[int, pd.Series, object]]] = defaultdict(list)
    for orig_i, pick, camp_region, ex_rank in pending:
        by_reg[camp_region].append((orig_i, pick, ex_rank))

    resolved_rows: list[pd.Series] = []
    for camp_region in sorted(by_reg.keys(), key=str):
        grp = by_reg[camp_region]
        ranks = [x[2] for x in grp]
        use_explicit = bool(
            ranks
            and all(pd.notna(x) for x in ranks)
            and all(1 <= int(x) <= n for x in ranks)
        )
        if use_explicit:
            grp = sorted(grp, key=lambda x: (int(x[2]), x[0]))
        else:
            grp = sorted(grp, key=lambda x: x[0])
        if len(grp) > n:
            warnings.append(
                f"권역 [{camp_region}]: {len(grp)}곳 중 상한 {n}곳만 반영했습니다(파일 위에서부터)."
            )
        grp = grp[:n]
        for seq, (_, pick, ex_rank) in enumerate(grp, start=1):
            pick = pick.copy()
            pick["campaign_region"] = camp_region
            pick["campaign_rank"] = int(ex_rank) if use_explicit else seq
            resolved_rows.append(pick)

    if not resolved_rows:
        return pd.DataFrame(), warnings, errors
    out = pd.DataFrame(resolved_rows).reset_index(drop=True)
    out = out.sort_values(["campaign_region", "campaign_rank", "name"]).reset_index(drop=True)
    return out, warnings, errors


def build_campaign_execution_table(
    selected_stores: pd.DataFrame,
    schools: pd.DataFrame,
    n_near: int,
) -> pd.DataFrame:
    if selected_stores.empty or schools.empty:
        return pd.DataFrame(columns=CAMPAIGN_EXEC_COLUMNS)
    st_work = selected_stores.reset_index(drop=True)
    top_idx, top_dist, sub = _cached_batch_neighbors(st_work, schools, n_near)
    rows: list[dict[str, object]] = []
    for i, strow in st_work.iterrows():
        for j in range(top_idx.shape[1]):
            ji = int(top_idx[i, j])
            dkm = float(top_dist[i, j]) if not np.isnan(top_dist[i, j]) else float("nan")
            if ji < 0 or np.isnan(dkm):
                continue
            s = sub.iloc[ji]
            row: dict[str, object] = {
                "권역": str(strow.get("campaign_region", "") or ""),
                "우선순위": int(strow.get("campaign_rank", 0) or 0),
                "매장명": str(strow.get("name", "") or ""),
                "학교명": simplify_school_name(s.get("name", "")),
                "학교유형": school_type_display_label(s.get("school_type", "")),
                "학교주소": str(s.get("address", "") or ""),
                "직선거리(km)": round(dkm, 3),
                "연락상태": "미착수",
                "메모": "",
            }
            for internal, kr in SCHOOL_CONTACT_COLUMNS:
                v = s.get(internal, "")
                row[kr] = str(v).strip() if pd.notna(v) else ""
            rows.append(row)
    if not rows:
        return pd.DataFrame(columns=CAMPAIGN_EXEC_COLUMNS)
    out = pd.DataFrame(rows)
    out = out.sort_values(
        ["권역", "우선순위", "매장명", "직선거리(km)"],
        ascending=[True, True, True, True],
    ).reset_index(drop=True)
    return out[CAMPAIGN_EXEC_COLUMNS]


def build_school_dedup_table(exec_df: pd.DataFrame) -> pd.DataFrame:
    if exec_df.empty:
        return pd.DataFrame(
            columns=[
                "학교명",
                "학교유형",
                "캠퍼스구분",
                "학교주소",
                "연관매장수",
                "최대직선거리(km)",
                "관련지역",
                "매장목록",
                "담당부서",
                "담당자명",
                "전화번호",
                "이메일",
            ]
        )
    rows: list[dict[str, object]] = []
    work = exec_df.copy()
    if "학교주소" not in work.columns:
        work["학교주소"] = ""
    work["학교명"] = work["학교명"].astype(str).str.strip()
    work["학교주소"] = work["학교주소"].astype(str).str.strip()
    for (school, school_addr), grp in work.groupby(["학교명", "학교주소"], sort=True):
        stores = sorted({str(x).strip() for x in grp["매장명"].tolist() if str(x).strip()})
        regions_col = "권역" if "권역" in grp.columns else ("campaign_region" if "campaign_region" in grp.columns else "")
        regions = sorted(
            {str(x).strip() for x in grp[regions_col].tolist() if str(x).strip()}
        ) if regions_col else []

        def _first_non_empty(series: pd.Series) -> str:
            for v in series.tolist():
                s = str(v).strip()
                if s:
                    return s
            return ""

        rows.append(
            {
                "학교명": str(school),
                "학교유형": _first_non_empty(grp["학교유형"]) if "학교유형" in grp.columns else "",
                "캠퍼스구분": _first_non_empty(grp["캠퍼스구분"]) if "캠퍼스구분" in grp.columns else "",
                "학교주소": str(school_addr),
                "연관매장수": len(stores),
                "최대직선거리(km)": round(float(pd.to_numeric(grp["직선거리(km)"], errors="coerce").max()), 3),
                "관련지역": " · ".join(regions[:8]) + (f" …외 {len(regions)-8}" if len(regions) > 8 else ""),
                "매장목록": " · ".join(stores[:12]) + (f" …외 {len(stores)-12}" if len(stores) > 12 else ""),
                "담당부서": _first_non_empty(grp["담당부서"]),
                "담당자명": _first_non_empty(grp["담당자명"]),
                "전화번호": _first_non_empty(grp["전화번호"]),
                "이메일": _first_non_empty(grp["이메일"]),
            }
        )
    out = pd.DataFrame(rows)
    return out.sort_values(["연관매장수", "최대직선거리(km)", "학교명"], ascending=[False, True, True]).reset_index(drop=True)


def build_priority_recommend_table(dedup_df: pd.DataFrame) -> pd.DataFrame:
    if dedup_df.empty:
        return pd.DataFrame(columns=["우선순위", "학교명", "추천점수", "근거", "연관매장수", "최소직선거리(km)", "이메일", "전화번호"])
    w = dedup_df.copy()
    stores_n = pd.to_numeric(w["연관매장수"], errors="coerce").fillna(0)
    dist_km = pd.to_numeric(w["최소직선거리(km)"], errors="coerce").fillna(99.0)
    has_email = w["이메일"].astype(str).str.strip().str.len() > 0
    has_phone = w["전화번호"].astype(str).str.strip().str.len() > 0

    score = (
        stores_n.clip(0, 6) * 10
        + (8 - dist_km.clip(0, 8)) * 5
        + has_email.astype(int) * 12
        + has_phone.astype(int) * 8
    )
    w["추천점수"] = score.round(1)
    w["근거"] = (
        "연관매장 "
        + stores_n.astype(int).astype(str)
        + "개 · 거리 "
        + dist_km.round(2).astype(str)
        + "km · 연락처 "
        + (has_email | has_phone).map({True: "있음", False: "부족"})
    )
    w = w.sort_values(["추천점수", "연관매장수", "최소직선거리(km)"], ascending=[False, False, True]).reset_index(drop=True)
    w["우선순위"] = np.arange(1, len(w) + 1, dtype=int)
    return w[["우선순위", "학교명", "추천점수", "근거", "연관매장수", "최소직선거리(km)", "이메일", "전화번호"]]



def kakao_map_html(
    center_lat: float,
    center_lng: float,
    app_key: str,
    points: list[dict],
    *,
    map_height_px: int = 280,
    map_level: int = 8,
    show_school_captions: bool = False,
) -> str:
    """학교는 기본 마커만(라벨 오버레이 생략) → 가볍게 렌더."""
    pts = json.dumps(points, ensure_ascii=False)
    key = html.escape(app_key, quote=True)
    sc = str(show_school_captions).lower()
    return f"""<!DOCTYPE html>
<html><head>
<meta charset="utf-8"/>
<meta http-equiv="Content-Security-Policy" content="upgrade-insecure-requests"/>
</head>
<body style="margin:0;font-family:{FONT_STACK_HTML};">
<div id="map" style="width:100%;height:{map_height_px}px;"></div>
<p id="map-err" style="display:none;margin:0;padding:10px 12px;font-size:13px;line-height:1.45;color:#b71c1c;background:#ffebee;border:1px solid #e57373;"></p>
<script>
function showMapErr(msg) {{
  var m = document.getElementById('map');
  var e = document.getElementById('map-err');
  if (m) m.style.display = 'none';
  if (e) {{ e.style.display = 'block'; e.textContent = msg; }}
}}
</script>
<script src="https://dapi.kakao.com/v2/maps/sdk.js?appkey={key}&autoload=false"
  onerror="showMapErr('카카오 지도 스크립트를 불러오지 못했습니다. 네트워크·광고차단·카카오 개발자(플랫폼) 웹 도메인 등록을 확인하세요. 로컬이면 http://localhost:포트 를 앱 플랫폼에 등록해야 합니다.')"></script>
<script>
var pts = {pts};
var showSchoolCaptions = {sc};
function pinDataUrl(hex) {{
  var svg = '<svg xmlns="http://www.w3.org/2000/svg" width="28" height="36" viewBox="0 0 36 48">' +
    '<path fill="' + hex + '" d="M18 2C11 2 6 7 6 14c0 10 12 24 12 24s12-14 12-24c0-7-5-12-12-12z"/>' +
    '<circle fill="#fff" cx="18" cy="14" r="4"/></svg>';
  return 'data:image/svg+xml;charset=UTF-8,' + encodeURIComponent(svg);
}}
if (typeof kakao === 'undefined' || !kakao.maps) {{
  showMapErr('카카오 지도 객체를 불러오지 못했습니다. JavaScript 앱키(kakao_js_key)와 developers.kakao.com → 플랫폼 → Web 사이트 도메인(예: http://localhost:8501 , 현재 쓰는 포트 포함)을 확인하세요.');
}} else {{
  kakao.maps.load(function() {{
    try {{
      var defC = new kakao.maps.LatLng({center_lat}, {center_lng});
      var map = new kakao.maps.Map(document.getElementById('map'), {{ center: defC, level: {map_level} }});
      var sz = new kakao.maps.Size(28, 36);
      var off = new kakao.maps.Point(14, 36);
      var szH = new kakao.maps.Size(34, 44);
      var offH = new kakao.maps.Point(17, 44);
      var szStoreMain = new kakao.maps.Size(38, 50);
      var offStoreMain = new kakao.maps.Point(19, 50);
      var imgStore = new kakao.maps.MarkerImage(pinDataUrl('{STARBUCKS_GREEN}'), sz, {{ offset: off }});
      var imgStoreMain = new kakao.maps.MarkerImage(pinDataUrl('{STARBUCKS_GREEN}'), szStoreMain, {{ offset: offStoreMain }});
      var imgStoreH = new kakao.maps.MarkerImage(pinDataUrl('{STARBUCKS_GREEN}'), szH, {{ offset: offH }});
      var imgNearStore = new kakao.maps.MarkerImage(pinDataUrl('rgba(0,131,72,0.45)'), sz, {{ offset: off }});
      var imgNearStoreH = new kakao.maps.MarkerImage(pinDataUrl('rgba(0,131,72,0.45)'), szH, {{ offset: offH }});
      var imgSchool = new kakao.maps.MarkerImage(pinDataUrl('#1565C0'), sz, {{ offset: off }});
      var imgSchoolH = new kakao.maps.MarkerImage(pinDataUrl('#1565C0'), szH, {{ offset: offH }});
      var bounds = new kakao.maps.LatLngBounds();
      var schoolLabelGroups = new Map();
      pts.sort(function(a, b) {{
        return ((a.kind === 'store') ? 1 : 0) - ((b.kind === 'store') ? 1 : 0);
      }});
      pts.forEach(function(p) {{
        var pos = new kakao.maps.LatLng(p.lat, p.lng);
        bounds.extend(pos);
        var kind = p.kind || 'school';
        var isHighlight = !!p.highlight;
        var mimg = imgSchool;
        var mimgH = imgSchoolH;
        if (kind === 'store') mimg = imgStore;
        else if (kind === 'near_store') mimg = imgNearStore;
        if (kind === 'store') mimgH = imgStoreH;
        else if (kind === 'near_store') mimgH = imgNearStoreH;
        if (kind === 'store') mimg = imgStoreMain;
        if (kind !== 'store' && isHighlight) mimg = mimgH;
        var marker = new kakao.maps.Marker({{ position: pos, image: mimg }});
        marker.setMap(map);
        var hoverOverlay = null;
        var cap = (p.caption && String(p.caption)) || (p.label && String(p.label).replace(/^매장:\\s*/, '')) || '';
        if (kind === 'near_store' && cap) {{
          var hwrap = document.createElement('div');
          hwrap.style.cssText = 'max-width:min(240px,80vw);padding:2px 6px;background:rgba(255,255,255,0.96);border:1px solid {STARBUCKS_GREEN};border-radius:3px;font-size:11px;font-weight:600;margin-bottom:30px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;';
          hwrap.style.pointerEvents = 'none';
          hwrap.appendChild(document.createTextNode(cap));
          hwrap.title = cap;
          hoverOverlay = new kakao.maps.CustomOverlay({{
            position: pos,
            content: hwrap,
            yAnchor: 1,
            xAnchor: 0.5,
            zIndex: 45
          }});
        }}
        kakao.maps.event.addListener(marker, 'mouseover', function() {{
          marker.setImage(mimgH);
          if (hoverOverlay) hoverOverlay.setMap(map);
        }});
        kakao.maps.event.addListener(marker, 'mouseout', function() {{
          marker.setImage(mimg);
          if (hoverOverlay) hoverOverlay.setMap(null);
        }});
        kakao.maps.event.addListener(marker, 'click', function() {{
          // 인근매장 라벨은 hover 전용: 클릭 시 남지 않도록 즉시 숨김
          if (hoverOverlay) hoverOverlay.setMap(null);
        }});
        if (kind === 'school' && cap && showSchoolCaptions) {{
          var k = String(p.lat).slice(0, 10) + ',' + String(p.lng).slice(0, 11);
          if (!schoolLabelGroups.has(k)) schoolLabelGroups.set(k, {{ pos: pos, caps: [] }});
          schoolLabelGroups.get(k).caps.push(cap);
        }} else if (cap && kind === 'store') {{
          var wrap = document.createElement('div');
          var capMb = (kind === 'store') ? 54 : (isHighlight ? 40 : 30);
          wrap.style.cssText = 'max-width:min(280px,85vw);padding:2px 6px;background:rgba(255,255,255,0.94);border:1px solid #444;border-radius:3px;font-size:11px;font-weight:600;margin-bottom:' + capMb + 'px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;';
          wrap.style.borderColor = kind === 'store' ? '{STARBUCKS_GREEN}' : '#1565C0';
          wrap.appendChild(document.createTextNode(cap));
          wrap.title = cap;
          new kakao.maps.CustomOverlay({{ position: pos, content: wrap, yAnchor: 1, xAnchor: 0.5, zIndex: kind === 'store' ? 120 : (isHighlight ? 60 : 10) }}).setMap(map);
        }}
      }});
      schoolLabelGroups.forEach(function(g) {{
        var caps = Array.from(new Set((g.caps || []).filter(Boolean)));
        if (!caps.length) return;
        var stack = document.createElement('div');
        stack.style.cssText = 'display:flex;flex-direction:column;gap:2px;align-items:center;margin-bottom:34px;';
        caps.forEach(function(txt) {{
          var box = document.createElement('div');
          box.style.cssText = 'max-width:min(260px,80vw);padding:1px 5px;background:rgba(255,255,255,0.94);border:1px solid #1565C0;border-radius:3px;font-size:11px;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;';
          box.appendChild(document.createTextNode(txt));
          box.title = txt;
          stack.appendChild(box);
        }});
        new kakao.maps.CustomOverlay({{ position: g.pos, content: stack, yAnchor: 1, xAnchor: 0.5, zIndex: 25 }}).setMap(map);
      }});
      if (pts.length >= 2) {{
        map.setBounds(bounds, 40, 40, 40, 40);
      }} else if (pts.length === 1) {{
        map.setCenter(new kakao.maps.LatLng(pts[0].lat, pts[0].lng));
        map.setLevel(7);
      }}
    }} catch (err) {{
      showMapErr('지도 표시 오류: ' + (err && err.message ? err.message : String(err)));
    }}
  }});
}}
</script>
</body></html>"""


def render_kakao_map(
    slat: float,
    slon: float,
    store_caption: str,
    app_key: str,
    *,
    nearby_top5: pd.DataFrame | None = None,
    nearby_stores: pd.DataFrame | None = None,
    highlight_school_key: str = "",
    map_height_px: int = 280,
    show_school_captions: bool = False,
) -> None:
    """인근 학교는 nearby_top5의 _geo_lat/_geo_lng만 사용. 강조는 _school_key(원본 이름+주소)로만 매칭."""
    h_key = str(highlight_school_key or "").strip()
    pts: list[dict] = []
    if nearby_top5 is not None and not nearby_top5.empty and "_geo_lat" in nearby_top5.columns:
        for _, r in nearby_top5.iterrows():
            if pd.isna(r.get("_geo_lat")) or pd.isna(r.get("_geo_lng")):
                continue
            school_name = str(r["학교명"]).strip()
            sk = str(r.get("_school_key", "")).strip()
            is_hl = bool(h_key and sk and sk == h_key)
            pts.append(
                {
                    "lat": float(r["_geo_lat"]),
                    "lng": float(r["_geo_lng"]),
                    "kind": "school",
                    "caption": school_name,
                    "label": school_name,
                    "highlight": is_hl,
                }
            )
    pts.append(
        {
            "lat": slat,
            "lng": slon,
            "kind": "store",
            "caption": store_caption,
            "label": f"매장: {store_caption}",
            "highlight": True,
        }
    )
    if nearby_stores is not None and not nearby_stores.empty and "_geo_lat" in nearby_stores.columns:
        for _, r in nearby_stores.iterrows():
            if pd.isna(r.get("_geo_lat")) or pd.isna(r.get("_geo_lng")):
                continue
            cap = str(r.get("매장명", "")).strip()
            pts.append(
                {
                    "lat": float(r["_geo_lat"]),
                    "lng": float(r["_geo_lng"]),
                    "kind": "near_store",
                    "caption": cap,
                    "label": "",
                    "highlight": False,
                }
            )
    h = map_height_px + 6
    components.html(
        kakao_map_html(
            slat,
            slon,
            app_key,
            pts,
            map_height_px=map_height_px,
            show_school_captions=show_school_captions,
        ),
        width=1100,
        height=h,
        scrolling=True,
    )


def _table_n_near_slider_and_value() -> int:
    """매장별 근접학교 요약·학교별 근접매장 요약 탭에서만 사용."""
    st.slider(
        "표·요약: 매장당 인근 학교 수 (1~5)",
        min_value=1,
        max_value=5,
        key="table_n_near",
        help="«매장별 근접학교 요약»와 «학교별 근접매장 요약»에만 적용됩니다. "
        "값이 작을수록 매장마다 보는 인근 학교 후보가 줄어들어, 학교별 요약 표에 나오는 학교 수도 함께 줄어듭니다. "
        "«매장·지도»·산학연계와는 별도입니다.",
    )
    return max(1, min(5, int(st.session_state["table_n_near"])))


def _require_app_access_password() -> None:
    """st.secrets['app_access_password']가 비어 있지 않으면 앱 전체를 비밀번호로 보호합니다.

    로컬에서 secrets를 쓰지 않거나 비밀번호를 비우면 기존처럼 바로 열립니다.
    배포(Streamlit Cloud 등)에서는 Secrets에 동일 키로 설정하세요.
    """
    try:
        expected = str(st.secrets.get("app_access_password", "") or "").strip()
    except Exception:
        expected = ""
    if not expected:
        return
    if st.session_state.get("_app_access_ok"):
        return
    st.markdown(
        f'<link rel="stylesheet" href="{PRETENDARD_CSS}" />'
        f"<style>.stApp {{ font-family: {FONT_STACK} !important; }}</style>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<p style="font-size:1.35rem;font-weight:800;color:{STARBUCKS_GREEN};margin:1rem 0 0.35rem 0;">'
        f"{html.escape(APP_TITLE)}</p>"
        "<p style=\"color:#3f3f46;font-size:1.02rem;\">내부용 페이지입니다. 배포 시 설정한 접속 비밀번호를 입력하세요.</p>",
        unsafe_allow_html=True,
    )
    with st.form("app_access_gate"):
        pw = st.text_input("비밀번호", type="password", autocomplete="current-password")
        submitted = st.form_submit_button("들어가기")
    if submitted:
        eu = pw.encode("utf-8")
        er = expected.encode("utf-8")
        ok = len(eu) == len(er) and hmac.compare_digest(eu, er)
        if ok:
            st.session_state["_app_access_ok"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    st.stop()


def main() -> None:
    st.set_page_config(page_title=PAGE_TITLE, layout="wide")
    _require_app_access_password()
    st.markdown(
        f'<link rel="stylesheet" href="{PRETENDARD_CSS}" />'
        "<style>"
        f".stApp, .stApp header, .stApp [data-testid=\"stAppViewContainer\"], "
        f".stApp .stMarkdown, .stApp label, .stApp p, .stApp li, .stApp td, .stApp th, "
        f".stApp input, .stApp textarea, .stApp select, "
        f".stApp [data-testid=\"stMarkdownContainer\"], .stApp .stCaption {{ "
        f"font-family: {FONT_STACK} !important; "
        f"}}"
        ".stApp p, .stApp li, .stApp label {font-size:1.10rem!important;}"
        ".stApp .stCaption, .stApp [data-testid=\"stCaptionContainer\"] p {font-size:0.96rem!important;}"
        ".stApp input, .stApp textarea, .stApp [data-baseweb=\"select\"] {font-size:1.07rem!important;}"
        f".stApp button, .stApp [data-baseweb] {{ font-family: {FONT_STACK}, sans-serif !important; }}"
        f".stApp a, .stApp a:visited {{ color: {STARBUCKS_GREEN} !important; }}"
        f".stApp [data-testid=\"stSlider\"] [role=\"slider\"] {{ background-color: {STARBUCKS_GREEN}; }}"
        f"[data-testid=\"stAppViewContainer\"], [data-testid=\"stMainBlockContainer\"] "
        f"{{ background: #fff !important; }}"
        ".stApp { -webkit-font-smoothing: antialiased; }"
        "html { color-scheme: only light; }"
        "section[data-testid=\"stSidebar\"] [data-testid=\"stMetricValue\"] {"
        "line-height:1.15!important;margin-bottom:0!important;padding-bottom:0!important;}"
        "section[data-testid=\"stSidebar\"] [data-testid=\"stMetricLabel\"] {"
        "margin-bottom:0.1rem!important;}"
        "section[data-testid=\"stSidebar\"] [data-testid=\"metric-container\"] {"
        "gap:0.1rem!important;}"
        "section[data-testid=\"stSidebar\"] [data-testid=\"stCaptionContainer\"] p{"
        "font-size:0.93rem!important;line-height:1.42!important;}"
        ".metric-ref-date{font-size:0.72rem;line-height:1.15;color:rgba(49,51,63,0.65);"
        "white-space:nowrap;margin:-0.6rem 0 0.2rem 0!important;padding:0!important;}"
        "div[data-testid=\"stSegmentedControl\"] [role=\"tablist\"]{gap:0.55rem!important;}"
        "div[data-testid=\"stSegmentedControl\"] [role=\"tab\"]{"
        "min-height:64px!important;padding:0.8rem 1.35rem!important;border-radius:12px!important;"
        "border:1px solid #b8c6cf!important;background:#eef2f5!important;color:#1f2a33!important;"
        "font-weight:900!important;font-size:1.18rem!important;letter-spacing:-0.01em!important;line-height:1.15!important;}"
        "div[data-testid=\"stSegmentedControl\"] [role=\"tab\"][aria-selected=\"true\"]{"
        "background:#0f172a!important;color:#ffffff!important;border-color:#0b1220!important;"
        "box-shadow:0 0 0 2px rgba(15,23,42,0.45), 0 8px 16px rgba(2,6,23,0.35)!important;"
        "transform:translateY(-1px)!important;}"
        "div[data-testid=\"stSegmentedControl\"] [role=\"tab\"]:last-child{min-width:255px!important;}"
        "div[data-testid=\"stSegmentedControl\"] [role=\"tab\"]:not(:last-child){min-width:160px!important;}"
        "div[data-testid=\"stDownloadButton\"] > button{min-height:30px!important;padding:0.18rem 0.56rem!important;"
        "font-size:0.84rem!important;line-height:1.08!important;border-radius:8px!important;}"
        "div[data-testid=\"stDownloadButton\"]{margin-top:0.08rem!important;margin-bottom:0.16rem!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"gridcell\"]{white-space:normal!important;line-height:1.35!important;padding-top:0.5rem!important;padding-bottom:0.5rem!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"columnheader\"]{white-space:normal!important;line-height:1.2!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"gridcell\"] *{white-space:normal!important;overflow:visible!important;text-overflow:clip!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"columnheader\"] *{white-space:normal!important;overflow:visible!important;text-overflow:clip!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"gridcell\"]{text-align:left!important;justify-content:flex-start!important;}"
        "div[data-testid=\"stDataFrame\"] [role=\"columnheader\"]{text-align:left!important;justify-content:flex-start!important;}"
        f"div[data-testid=\"stButton\"] > button[kind=\"primary\"]{{background:{STARBUCKS_GREEN}!important;"
        f"color:#fff!important;border:1px solid {STARBUCKS_GREEN}!important;font-weight:800!important;"
        "box-shadow:0 4px 10px rgba(0,131,72,0.26)!important;}"
        "div[data-testid=\"stButton\"] > button[kind=\"primary\"]:hover{filter:brightness(0.95)!important;}"
        f"</style>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f'''<div style="display:flex;align-items:center;gap:18px;flex-wrap:wrap;margin:0 0 4px 0;padding:4px 4px 10px 4px;border-bottom:2px solid {STARBUCKS_GREEN};background:linear-gradient(180deg,rgba(0,112,74,0.06) 0%,rgba(0,112,74,0) 55%);font-family:{FONT_STACK};">
  <h1 style="margin:0;padding:0;font-size:1.95rem;font-weight:800;color:{STARBUCKS_GREEN};letter-spacing:-0.03em;line-height:1.15;border:none;font-family:{FONT_STACK};">{html.escape(APP_TITLE)}</h1>
</div>''',
        unsafe_allow_html=True,
    )

    credit_line = ""
    last_updated = ""
    try:
        if "app_credit" in st.secrets:
            credit_line = str(st.secrets["app_credit"]).strip()
        if "app_last_updated" in st.secrets:
            last_updated = str(st.secrets["app_last_updated"]).strip()
    except Exception:
        pass
    if credit_line or last_updated:
        _cred_style = "font-size:0.98rem;font-weight:450;color:#3f3f3f;"
        if credit_line and last_updated:
            meta = f"※ 데이터 갱신: {last_updated}"
            inner = (
                f'<span style="{_cred_style}">{html.escape(credit_line)}</span>'
                f'<span style="{_cred_style} margin-left:2rem;">{html.escape(meta)}</span>'
            )
        elif credit_line:
            inner = f'<span style="{_cred_style}">{html.escape(credit_line)}</span>'
        else:
            meta = f"※ 데이터 갱신: {last_updated}"
            inner = f'<span style="{_cred_style}">{html.escape(meta)}</span>'
        st.markdown(f'<p style="line-height:1.5;margin:0 0 0.6rem 0;">{inner}</p>', unsafe_allow_html=True)

    if not DATA.is_file():
        st.error(
            "processed_data.csv가 없습니다. import_excel_to_csv.py 후 "
            "`python 2_geocode_kakao_all.py` 또는 `python 2_geocode.py` 를 실행하세요."
        )
        st.stop()

    _mtime = int(DATA.stat().st_mtime_ns)
    df = load_data(str(DATA), _mtime)
    stores_from_csv = _ensure_store_meta(df[df["entity_type"] == "store"].copy())
    schools = df[df["entity_type"] == "school"].copy()
    if "school_type" in schools.columns:
        schools["school_type"] = schools["school_type"].map(normalize_school_type_value)

    if schools.empty:
        st.error("학교 데이터가 비어 있습니다. processed_data.csv를 확인하세요.")
        st.stop()

    try:
        kakao_secret = str(st.secrets.get("kakao_js_key", "") or "").strip()
    except Exception:
        kakao_secret = ""

    store_ref = ""
    school_ref = ""
    try:
        if "store_reference_date" in st.secrets:
            store_ref = str(st.secrets["store_reference_date"]).strip()
        if "school_reference_date" in st.secrets:
            school_ref = str(st.secrets["school_reference_date"]).strip()
    except Exception:
        pass

    src = "csv"
    upload_file_hash = ""
    st.caption(
        "내부 분석용 · 데이터 `processed_data.csv` · 좌측 필터는 전 탭에 동일 적용 · "
        "엑셀 업로드(매장 특성·산학연계)는 각각 해당 영역에서만."
    )
    stores_base = apply_store_tags(stores_from_csv.copy())
    if stores_base.empty:
        st.error("processed_data.csv에 매장 행이 없습니다.")
        st.stop()

    stores_base = _ensure_store_meta(stores_base.reset_index(drop=True))

    team_vals = sorted(
        {str(x).strip() for x in stores_base["ops_team"].tolist() if str(x).strip()},
        key=_team_sort_key,
    )
    region_vals = sorted(
        {str(x).strip() for x in stores_base["store_region"].tolist() if str(x).strip()}
    )

    with st.sidebar:
        m1, m2 = st.columns(2)
        with m1:
            st.metric("매장", f"{len(stores_base):,}")
            st.markdown(
                '<p class="metric-ref-date">26.3.31 기준</p>',
                unsafe_allow_html=True,
            )
        with m2:
            st.metric("학교", f"{len(schools):,}")
            st.markdown(
                '<p class="metric-ref-date">26.4.1 기준</p>',
                unsafe_allow_html=True,
            )
        with st.expander("표시·세션 안내", expanded=False):
            _mb = _MAX_XLSX_UPLOAD_BYTES // (1024 * 1024)
            st.caption(
                "새로고침·재접속 시 일부 선택·업로드 상태가 초기화될 수 있습니다. 중요 표는 다운로드로 보관하세요."
            )
            st.caption(
                "`산학연계(Summary)`는 템플릿 업로드 후 「작업 시작」을 눌러야 결과가 생성됩니다."
            )
            st.caption(f"엑셀 업로드는 파일당 {_mb}MB 이하만 허용됩니다.")
        with st.expander("매장 특성 설정", expanded=False):
            tag_counts = {c: int((stores_base[c] == "Y").sum()) if c in stores_base.columns else 0 for c in STORE_TAG_COLUMNS}
            st.caption("엑셀로 특성(Y/N) 입력 후 업로드하면 저장됩니다.")
            st.caption(
                " · ".join([f"{k} {v:,}" for k, v in tag_counts.items()])
            )
            tdf = build_store_tags_template(stores_base)
            tbuf = BytesIO()
            with pd.ExcelWriter(tbuf, engine="openpyxl") as w:
                tdf.to_excel(w, index=False, sheet_name="매장특성")
            tbuf.seek(0)
            st.download_button(
                "매장 특성 템플릿 다운로드",
                data=tbuf.read(),
                file_name="store_tags_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_store_tags_template",
            )
            if st.button("특성 초기화(모두 N)", key="reset_store_tags_btn"):
                if STORE_TAGS_FILE.is_file():
                    STORE_TAGS_FILE.unlink(missing_ok=True)
                st.success("저장된 매장 특성을 초기화했습니다.")
                st.rerun()
            up_tags = st.file_uploader(
                "매장 특성 업로드 (.xlsx)",
                type=["xlsx"],
                key="store_tags_upload_xlsx",
                help=f".xlsx만 가능 · 파일당 최대 {_MAX_XLSX_UPLOAD_BYTES // (1024 * 1024)}MB",
            )
            if up_tags is not None and st.button("매장 특성 업로드 반영", key="apply_store_tags_upload_btn", type="primary"):
                _sz = _xlsx_upload_size_error(up_tags)
                if _sz:
                    st.error(_sz)
                else:
                    try:
                        merged = apply_store_tags_upload(up_tags.getvalue(), stores_base)
                        save_store_tags(merged)
                        # 특성 업로드 직후 요약 표가 이전 캐시를 보는 경우를 방지
                        st.cache_data.clear()
                        st.success("매장 특성 업로드를 반영했습니다.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"매장 특성 업로드 처리 실패: {e}")
        with st.expander("매장 특성 필터", expanded=False):
            pick_store_tags_and = st.multiselect(
                "AND 조건(선택 특성만 Y)",
                options=list(STORE_TAG_COLUMNS),
                default=[],
                key="store_tag_filter_keys_and",
            )
            pick_store_tags_or = st.multiselect(
                "OR 조건(하나 이상 Y)",
                options=list(STORE_TAG_COLUMNS),
                default=[],
                key="store_tag_filter_keys_or",
            )
        with st.expander("학교 유형 필터", expanded=False):
            all_school_filter_keys = [fk for fk, _ in SCHOOL_FILTER_DEF]
            for fk in all_school_filter_keys:
                st.session_state.setdefault(f"sf_{fk}", True)

            _sf_c1, _sf_c2 = st.columns(2)
            with _sf_c1:
                if st.button("전체선택", key="sf_select_all_btn", use_container_width=True):
                    for fk in all_school_filter_keys:
                        st.session_state[f"sf_{fk}"] = True
                    st.rerun()
            with _sf_c2:
                if st.button("초기화", key="sf_reset_btn", use_container_width=True):
                    for fk in all_school_filter_keys:
                        st.session_state[f"sf_{fk}"] = True
                    st.rerun()

            selected_keys: set[str] = set()
            for group_name, pairs in SCHOOL_FILTER_GROUPS:
                st.caption(group_name)
                for fk, flabel in pairs:
                    if st.checkbox(flabel, value=True, key=f"sf_{fk}"):
                        selected_keys.add(fk)
            if not selected_keys:
                schools_use = schools.copy()
            else:
                schools_use = filter_schools_by_keys(schools, selected_keys)
            counts = school_key_counts(schools_use)
            tail = " · ".join(
                [f"{lbl} {counts[fk]}" for fk, lbl in SCHOOL_FILTER_DEF if counts.get(fk, 0)]
            )
            st.caption(
                f"적용 학교 **{len(schools_use)}**/{len(schools)}개교" + (f" · {tail}" if tail else "")
            )
            selected_labels = [lbl for fk, lbl in SCHOOL_FILTER_DEF if fk in selected_keys]
            st.caption(
                "현재 선택: " + (", ".join(selected_labels) if selected_labels else "(없음: 전체 학교 적용)")
            )
        with st.expander("운영팀·권역", expanded=False):
            st.caption("엑셀·CSV에 운영팀·권역 열이 있을 때만 줄일 수 있습니다.")
            pick_teams = (
                st.multiselect("운영팀", options=team_vals, default=team_vals)
                if team_vals
                else []
            )
            pick_regions = (
                st.multiselect("권역", options=region_vals, default=region_vals)
                if region_vals
                else []
            )
        with st.expander("지도 API 키", expanded=not bool(kakao_secret)):
            kakao_key_input = st.text_input(
                "JavaScript",
                value=kakao_secret,
                type="password",
                label_visibility="collapsed",
                placeholder="secrets.toml 또는 입력",
            )
    key = (kakao_key_input or "").strip() or kakao_secret
    if not key:
        st.caption(
            "지도: 사이드바 «지도 API 키»에 Kakao **JavaScript** 키를 넣으면 «매장·지도»에서 카카오 지도가 표시됩니다."
        )

    # 표·요약 탭(매장별 근접학교·학교별 근접매장) 전용 슬라이더 값. 산학연계 campaign_n_near와 별도.
    if "table_n_near" not in st.session_state:
        st.session_state["table_n_near"] = 5

    filt_ok = True
    mask = pd.Series(True, index=stores_base.index)
    if team_vals:
        if not pick_teams:
            filt_ok = False
        else:
            mask &= stores_base["ops_team"].astype(str).str.strip().isin(pick_teams)
    if region_vals:
        if not pick_regions:
            filt_ok = False
        else:
            mask &= stores_base["store_region"].astype(str).str.strip().isin(pick_regions)
    if pick_store_tags_and or pick_store_tags_or:
        mt = np.ones(len(stores_base), dtype=bool)
        if pick_store_tags_and:
            for tc in STORE_TAG_COLUMNS:
                col_is_y = stores_base[tc].map(_norm_tag_val).eq("Y").to_numpy()
                if tc in pick_store_tags_and:
                    mt &= col_is_y
                else:
                    mt &= ~col_is_y
        if pick_store_tags_or:
            mor = np.zeros(len(stores_base), dtype=bool)
            for tc in pick_store_tags_or:
                mor |= stores_base[tc].map(_norm_tag_val).eq("Y").to_numpy()
            mt &= mor
        mask &= mt
    if not filt_ok and (team_vals or region_vals):
        stores_view = stores_base.iloc[0:0].copy()
        st.warning("운영팀·권역을 각각 한 가지 이상 선택하세요.")
    else:
        stores_view = stores_base[mask].copy()

    sum1, sum2 = st.columns(2)
    with sum1:
        st.caption(f"선택된 매장 {len(stores_view):,}곳(26.3.31기준)")
    with sum2:
        st.caption(f"학교 풀 {len(schools_use):,}교(26.4.1기준)")
    _scope_bits: list[str] = [
        f"매장 {len(stores_view):,}곳",
        f"학교 {len(schools_use):,}교",
    ]
    if team_vals:
        _scope_bits.append(f"운영팀 {len(pick_teams)}개")
    if region_vals:
        _scope_bits.append(f"권역 {len(pick_regions)}개")
    if pick_store_tags_and or pick_store_tags_or:
        _scope_bits.append("매장 특성 필터 ON")
    _scope_html = (
        '<p style="margin:0.4rem 0 0.65rem 0;padding:0.55rem 0.85rem;border-left:4px solid '
        + STARBUCKS_GREEN
        + ';background:rgba(0,112,74,0.07);border-radius:0 8px 8px 0;font-size:0.98rem;color:#0f172a;line-height:1.45;">'
        '<strong style="color:#0f172a;">적용 범위</strong> · '
        + html.escape(" · ".join(_scope_bits))
        + "</p>"
    )
    st.markdown(_scope_html, unsafe_allow_html=True)

    _main_tab_labels = ["매장·지도", "매장별 근접학교 요약", "학교별 근접매장 요약", "산학연계(Summary)"]
    _legacy_tab = "매장별 근접 5개교 표"
    _new_near_tab = "매장별 근접학교 요약"
    if st.session_state.get("main_area_tab") == _legacy_tab:
        st.session_state["main_area_tab"] = _new_near_tab
    if st.session_state.get("_main_area_tab_last") == _legacy_tab:
        st.session_state["_main_area_tab_last"] = _new_near_tab
    _legacy_near_tab_old_name = "매장별 근접학교 표"
    if st.session_state.get("main_area_tab") == _legacy_near_tab_old_name:
        st.session_state["main_area_tab"] = _new_near_tab
    if st.session_state.get("_main_area_tab_last") == _legacy_near_tab_old_name:
        st.session_state["_main_area_tab_last"] = _new_near_tab
    _legacy_school_tab = "학교·팀 요약"
    _legacy_school_tab_v2 = "학교별 요약"
    _new_school_tab = "학교별 근접매장 요약"
    if st.session_state.get("main_area_tab") == _legacy_school_tab:
        st.session_state["main_area_tab"] = _new_school_tab
    if st.session_state.get("_main_area_tab_last") == _legacy_school_tab:
        st.session_state["_main_area_tab_last"] = _new_school_tab
    if st.session_state.get("main_area_tab") == _legacy_school_tab_v2:
        st.session_state["main_area_tab"] = _new_school_tab
    if st.session_state.get("_main_area_tab_last") == _legacy_school_tab_v2:
        st.session_state["_main_area_tab_last"] = _new_school_tab
    # 필터 변경 등으로 재실행될 때 탭이 첫 탭으로 튀지 않도록 마지막 탭을 고정 보존
    _last_tab = st.session_state.get("_main_area_tab_last")
    _widget_tab = st.session_state.get("main_area_tab")
    if _widget_tab in _main_tab_labels:
        _last_tab = _widget_tab
    if _last_tab not in _main_tab_labels:
        _last_tab = _main_tab_labels[0]
    st.session_state["_main_area_tab_last"] = _last_tab

    main_tab = st.segmented_control(
        "화면",
        options=_main_tab_labels,
        default=_last_tab,
        required=True,
        key="main_area_tab",
        label_visibility="collapsed",
        width="stretch",
    )
    if main_tab in _main_tab_labels:
        st.session_state["_main_area_tab_last"] = main_tab

    if main_tab == "매장·지도":
        map_near_n = 5
        if schools_use.empty:
            st.info("선택한 학교 구분에 해당하는 학교가 없습니다.")
        elif stores_view.empty:
            st.info("표시할 매장이 없습니다. 매장 목록·운영팀·권역을 확인하세요.")
        else:
            idx_key = "store_detail_idx"
            sig_key = "_store_pick_filter_sig"
            _suggest_ph = "— 추천에서 선택 —"
            st.markdown('<div style="height:14px;"></div>', unsafe_allow_html=True)
            st.info(
                "아래 목록에서 매장을 선택하면 인근 학교 표와 지도가 열립니다. "
                "처음에는 특정 매장이 잡혀 있지 않습니다."
            )
            _qcol, _xcol = st.columns([1, 0.14], gap="small")
            with _qcol:
                st.markdown(
                    '<p style="margin:0 0 0.2rem 0;font-size:1rem;font-weight:700;">'
                    '<span style="color:#0f172a;margin-right:6px;">■</span>'
                    "스타벅스 매장 검색"
                    "</p>",
                    unsafe_allow_html=True,
                )
                q = st.text_input(
                    "스타벅스 매장 검색",
                    placeholder="예: 강남(매장명 우선) · 서초(주소) · 운영팀 · 권역",
                    key="q1",
                    label_visibility="collapsed",
                )
            with _xcol:
                st.markdown('<div style="margin-top: 1.85rem;"></div>', unsafe_allow_html=True)
                if st.button("✕", key="clear_store_search", help="검색어 비우기 · 추천·표시 매장 초기화"):
                    st.session_state.pop("q1", None)
                    st.session_state.pop("store_suggest_pick", None)
                    st.session_state[idx_key] = STORE_PICK_NONE
                    st.session_state.pop(sig_key, None)
                    st.rerun()

            st.markdown(
                '<p style="margin:0.5rem 0 0.14rem 0;font-size:0.92rem;'
                'font-weight:700;color:#334155;">'
                '<span style="color:#0f172a;margin-right:6px;">■</span>'
                "추천 매장"
                "</p>"
                "",
                unsafe_allow_html=True,
            )
            spick_full = build_store_pick_frame(stores_view)
            sig = (
                src,
                upload_file_hash,
                frozenset(pick_teams) if pick_teams else frozenset(),
                frozenset(pick_regions) if pick_regions else frozenset(),
                q.strip(),
            )
            if st.session_state.get(sig_key) != sig:
                st.session_state[sig_key] = sig
                st.session_state[idx_key] = STORE_PICK_NONE
                if "store_suggest_pick" in st.session_state:
                    del st.session_state["store_suggest_pick"]
            sug = suggest_store_pick_labels(spick_full, q.strip())
            suggest_opts = [_suggest_ph] + sug
            if (
                "store_suggest_pick" not in st.session_state
                or st.session_state["store_suggest_pick"] not in suggest_opts
            ):
                st.session_state["store_suggest_pick"] = _suggest_ph
            _rec_map: dict[str, str] = dict(
                zip(
                    spick_full["pick_label"].astype(str),
                    spick_full["recommend_label"].astype(str),
                )
            )

            def _fmt_recommend_pick(opt: str) -> str:
                if opt == _suggest_ph:
                    return opt
                return _rec_map.get(opt, opt)

            st.selectbox(
                "recommend_pick",
                options=suggest_opts,
                key="store_suggest_pick",
                disabled=len(sug) == 0,
                label_visibility="collapsed",
                format_func=_fmt_recommend_pick,
            )
            pick_suggest = str(st.session_state.get("store_suggest_pick") or _suggest_ph)

            if pick_suggest != _suggest_ph:
                spick = spick_full.loc[spick_full["pick_label"] == pick_suggest].reset_index(drop=True)
            elif q.strip():
                qlow = q.strip().lower()
                mask_q = spick_full["pick_label"].str.lower().str.contains(qlow, regex=False)
                sub = spick_full.loc[mask_q].copy()
                pr = _store_search_match_priorities(sub, qlow)
                spick = (
                    sub.assign(_prio=pr)
                    .sort_values(["_prio", "pick_label"])
                    .drop(columns=["_prio"])
                    .reset_index(drop=True)
                )
            else:
                spick = spick_full

            if pick_suggest != _suggest_ph and not spick.empty and len(spick) == 1:
                st.session_state[idx_key] = 0

            if spick.empty:
                st.warning("검색 조건에 맞는 매장이 없습니다. 검색어를 지우거나 바꿔 보세요.")
            else:
                n_sp = len(spick)
                pick_options = [STORE_PICK_NONE] + list(range(n_sp))
                _cur = st.session_state.get(idx_key)
                if _cur not in pick_options:
                    st.session_state[idx_key] = STORE_PICK_NONE

                def _fmt_store_pick(j: int) -> str:
                    if j == STORE_PICK_NONE:
                        return STORE_PICK_LABEL_NONE
                    return str(spick.iloc[int(j)]["pick_label"])

                st.markdown(
                    '<p style="margin:0.1rem 0 0.35rem 0;font-size:1.02rem;font-weight:800;color:#0f172a;">'
                    "■ 표·지도를 볼 매장 (이름 · 주소 · 운영팀 · 권역)"
                    "</p>",
                    unsafe_allow_html=True,
                )
                pick_i = st.selectbox(
                    "표·지도를 볼 매장",
                    options=pick_options,
                    format_func=_fmt_store_pick,
                    key=idx_key,
                    label_visibility="collapsed",
                )

                if pick_i != STORE_PICK_NONE:
                    row = spick.iloc[int(pick_i)]

                    slat = float(row["latitude"])
                    slon = float(row["longitude"])
                    # 매장·지도의 인근학교는 항상 전체 학교 기준 최대 5개 고정
                    result = top5_for_store(slat, slon, schools, n=map_near_n)
                    nearby_stores = topn_nearby_stores(
                        slat,
                        slon,
                        stores_view,
                        n=5,
                        exclude_name=str(row.get("name", "")),
                        exclude_address=str(row.get("address", "")),
                    )
                    _ot = str(row.get("ops_team", "")).strip()
                    _rg = str(row.get("store_region", "")).strip()
                    bad_html: list[str] = []
                    if _ot:
                        bad_html.append(
                            "<span style=\"display:inline-block;background:#ecfdf5;color:#065f46;"
                            "padding:2px 8px;border:1px solid #6ee7b7;border-radius:3px;font-size:0.8rem;"
                            f"font-family:{FONT_MONO};margin:0 6px 6px 0;\">"
                            f"{html.escape(_ot)}</span>"
                        )
                    if _rg:
                        bad_html.append(
                            "<span style=\"display:inline-block;background:#eff6ff;color:#1e40af;"
                            "padding:2px 8px;border:1px solid #93c5fd;border-radius:3px;font-size:0.8rem;"
                            f"font-family:{FONT_MONO};margin:0 6px 6px 0;\">"
                            f"{html.escape(_rg)}</span>"
                        )
                    if bad_html:
                        st.markdown(" ".join(bad_html), unsafe_allow_html=True)
                    st.caption(str(row["address"]))
                    c_left, c_right = st.columns([1.05, 1.0], gap="medium")
                    with c_left:
                        st.markdown(f"##### 인근 학교 ({map_near_n})")
                        st.caption(
                            f"학교 좌표가 있는 전체 학교 풀에서, 선택 매장과의 "
                            f"직선거리(Haversine, km) 가까운 순 상위 {map_near_n}개입니다. "
                            "일정 거리(반경) 안만 보여 주는 방식이 아니며, «학교 유형 필터»는 이 표에는 적용되지 않습니다."
                        )
                        st.caption(
                            "※ 자차·대중교통 「추정(분)」은 외부 길찾기 API가 아니라, 같은 행의 직선거리(km)만으로 계산한 참고값입니다. "
                            + TRAVEL_TIME_HELP
                            + f" 유효거리(km) ≈ 직선×{_TRAVEL_ROAD_FACTOR}, 자차 {_TRAVEL_CAR_KMH}km/h·대중교통 {_TRAVEL_TRANSIT_KMH}km/h로 분을 환산합니다."
                        )
                        _show = result.drop(
                            columns=[c for c in ("_geo_lat", "_geo_lng", "_school_key") if c in result.columns],
                            errors="ignore",
                        )
                        render_table(_show, use_container_width=True)
                        st.markdown("##### 인근 매장 (5)")
                        st.caption(
                            "다른 매장 중 선택 매장과 직선거리(km) 가까운 순 상위 5곳입니다. "
                            "반경 km 조건 없이 가장 가까운 순으로만 고르며, 선택 매장 본인은 제외합니다."
                        )
                        st.caption(
                            "※ 인근 매장 표의 자차·대중교통 「추정(분)」도 위와 동일하게 직선거리(km) 기준 참고 추정입니다. "
                            + TRAVEL_TIME_HELP
                        )
                        _show_store = nearby_stores.drop(
                            columns=[c for c in ("_geo_lat", "_geo_lng") if c in nearby_stores.columns],
                            errors="ignore",
                        )
                        render_table(_show_store, use_container_width=True)
                    with c_right:
                        st.markdown("##### 지도")
                        st.caption(
                            "진한 초록 핀 = 선택 매장, 투명 초록 핀 = 인근 매장, 파란 핀 = 학교. "
                            "동명 학교는 아래에서 구분·강조합니다."
                        )
                        map_hl_key = ""
                        if not result.empty and "_school_key" in result.columns:
                            _ks = [str(x).strip() for x in result["_school_key"].tolist() if str(x).strip()]
                            _row_sig = f"{str(row.get('name', ''))}\x1e{str(row.get('address', ''))}"
                            _hl_widget_key = f"map_school_highlight_{hash(_row_sig) & 0x7FFFFFFF}"
                            map_hl_key = st.selectbox(
                                "지도에서 강조할 학교",
                                options=[""] + _ks,
                                format_func=lambda k: (
                                    "(강조 없음)"
                                    if not k
                                    else _near_school_table_label(
                                        result.loc[result["_school_key"].astype(str).str.strip() == k].iloc[0]
                                    )
                                ),
                                key=_hl_widget_key,
                            )
                        map_title = str(row["name"])[:40]
                        if key:
                            render_kakao_map(
                                slat,
                                slon,
                                map_title,
                                key,
                                nearby_top5=result,
                                nearby_stores=nearby_stores,
                                highlight_school_key=map_hl_key,
                                map_height_px=700,
                                show_school_captions=True,
                            )
                        else:
                            st.markdown(
                                f'<div style="min-height:700px;display:flex;align-items:center;justify-content:center;'
                                f"background:linear-gradient(165deg,#f8fafc 0%,#ecfdf5 55%,#f1f5f9 100%);"
                                f'border:1px dashed #cbd5e1;border-radius:10px;color:#334155;font-size:0.96rem;'
                                f'padding:1.25rem;text-align:center;line-height:1.55;">'
                                f'<div><span style="color:{STARBUCKS_GREEN};font-weight:800;font-size:1.02rem;">지도 미표시</span><br/>'
                                "좌측 «지도 API 키»에 Kakao <strong>JavaScript</strong> 키를 넣으면 이 영역에 카카오 지도가 열립니다."
                                "<br/><span style=\"font-size:0.84rem;color:#64748b;\">"
                                "플랫폼에 접속 주소(예: localhost:포트) 등록이 필요할 수 있습니다.</span></div></div>",
                                unsafe_allow_html=True,
                            )
    elif main_tab == "매장별 근접학교 요약":
        n_near_table = _table_n_near_slider_and_value()
        st.caption(
            f"행 = 매장 {len(stores_view):,} · 매장당 학교 {n_near_table} · 학교 풀 {len(schools_use):,}"
        )
        st.caption(
            "각 매장 행마다 가까운 학교 상위 N개(N=위 슬라이더)까지 열로 표시합니다. "
            "표의 행 수는 선택된 매장 수와 같습니다."
        )
        if schools_use.empty or stores_view.empty:
            st.info("표시할 데이터가 없습니다.")
        else:
            wide = build_all_stores_wide(stores_view, schools_use, n=n_near_table)
            _h_all_l, _h_all_r = st.columns([0.82, 0.18])
            with _h_all_l:
                st.markdown("##### 매장별 근접학교 요약")
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                wide.to_excel(w, index=False, sheet_name="전체")
            buf.seek(0)
            with _h_all_r:
                st.download_button(
                    "엑셀 다운로드",
                    data=buf.read(),
                    file_name="all_stores_top_schools.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_all_wide_xlsx",
                )
            render_wide_store_school_table(wide, use_container_width=True, height=480)

    elif main_tab == "학교별 근접매장 요약":
        n_near_table = _table_n_near_slider_and_value()
        if schools_use.empty or stores_view.empty:
            st.info("표시할 데이터가 없습니다.")
        else:
            active_tag_filters = []
            if pick_store_tags_and:
                active_tag_filters.append("AND: " + ", ".join(pick_store_tags_and))
            if pick_store_tags_or:
                active_tag_filters.append("OR: " + ", ".join(pick_store_tags_or))
            if active_tag_filters:
                st.caption(
                    f"특성 필터 반영 매장 {len(stores_view):,}곳 · " + " / ".join(active_tag_filters)
                )
            else:
                st.caption(
                    "특성 업로드만으로는 요약 수치가 바뀌지 않습니다. "
                    "좌측 `매장 특성 필터`(AND/OR)를 선택하면 이 화면에도 바로 반영됩니다."
                )
            st.markdown("##### 학교별 근접매장 요약 (근처 매장 많은 순)")
            sch_tbl = build_school_centric_table(stores_view, schools_use, n_near_table)
            st.caption(
                f"표 행 **{len(sch_tbl):,}**개 · "
                f"‘어느 매장의 상위 {n_near_table}개 인근 학교’에 한 번이라도 포함된 학교만 나옵니다. "
                f"학교 풀 **{len(schools_use):,}**교 전체와 행 수가 같지 않을 수 있습니다."
            )
            b2 = BytesIO()
            with pd.ExcelWriter(b2, engine="openpyxl") as w:
                sch_tbl.to_excel(w, index=False, sheet_name="학교중심")
            b2.seek(0)
            st.markdown(
                f"""<div style="color:rgba(49,51,63,0.78);font-size:0.96rem;line-height:1.55;margin:0.08rem 0 0.35rem 0;">
• 연관매장수: 각 매장 기준 가까운 학교 상위 {n_near_table}개에 해당 학교가 들어간 횟수(중복 매장 제외)<br>
• 직관적 해석: 값이 클수록 해당 학교 근처로 보는 매장이 많음<br>
• 최대직선거리: 연관 매장 중 해당 학교와 가장 먼 매장 거리<br>
<span style="color:{STARBUCKS_GREEN};font-weight:800;">• 예시: 진선여고 연관매장수가 높으면, 진선여고를 가까운 학교로 보는 매장이 많은 상태</span>
</div>""",
                unsafe_allow_html=True,
            )
            _sch_btn_l, _sch_btn_r = st.columns([0.82, 0.18])
            with _sch_btn_r:
                st.download_button(
                    "엑셀 다운로드",
                    data=b2.read(),
                    file_name="school_centric.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_school_centric_top_right",
                )
            render_school_centric_table(sch_tbl, use_container_width=True, height=360)
    elif main_tab == "산학연계(Summary)":
        st.caption(
            "절차: **빈 양식 다운로드** → 매장명 입력 → **업로드** → **「작업 시작」** (업로드만으로는 표가 생성되지 않습니다)."
        )
        st.caption(
            "좌측 «운영팀·권역», «매장 특성 필터»는 이 탭 산출에도 동일하게 적용됩니다. "
            "«학교 유형 필터»는 근접 학교·실행표에 반영됩니다."
        )
        stores_per_team = st.slider(
            "권역당 최대 취합 매장 수",
            min_value=1,
            max_value=5,
            value=3,
            step=1,
            key="campaign_stores_per_team",
            help="권역당 포함할 수 있는 매장 수 상한입니다. "
            "엑셀에는 «매장명»만 적어도 되며, 권역은 데이터 기준으로 붙고 우선순위는 파일 순서로 1,2,3… 부여됩니다.",
        )
        _campaign_n_near_val = int(st.session_state.get("campaign_n_near", 3))
        _campaign_n_near_val = max(1, min(5, _campaign_n_near_val))
        campaign_n_near = st.slider(
            "1개 매장 당 산학연계 학교 수",
            min_value=1,
            max_value=5,
            value=_campaign_n_near_val,
            step=1,
            key="campaign_n_near",
            help="업로드 후 하단 표(연락실행표/중복통합/우선연락추천)에 적용됩니다.",
        )
        st.markdown(f"##### 지역별 매장 취합 (권역당 최대 {stores_per_team}곳까지)")
        st.caption(
            "1) 빈 양식 다운로드 → 2) «매장명»만 적기(같은 권역은 위에서부터 순서대로 우선순위 부여) → 3) 업로드 · 선택적으로 권역·순위 열을 넣는 구 양식도 가능"
        )
        tmpl = build_campaign_template(blank_rows=max(32, stores_per_team * 10))
        tbuf = BytesIO()
        with pd.ExcelWriter(tbuf, engine="openpyxl") as w:
            tmpl.to_excel(w, index=False, sheet_name="입력템플릿")
            ws = w.book["입력템플릿"]
            thin = Side(style="thin", color="D0D7DE")
            header_fill = PatternFill(fill_type="solid", fgColor="EAF5F0")
            header_font = Font(name="Pretendard", size=11, bold=True, color="0F172A")
            body_font = Font(name="Pretendard", size=10, bold=False, color="111827")
            center = Alignment(horizontal="center", vertical="center")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            max_row = ws.max_row
            max_col = ws.max_column
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = center
                    cell.border = border
                    if r == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        cell.font = body_font

            width_map = {
                "매장명": 40,
                "권역": 24,
                "우선순위": 12,
            }
            for col_idx in range(1, max_col + 1):
                h = str(ws.cell(row=1, column=col_idx).value or "").strip()
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width_map.get(h, 22)
            ws.freeze_panes = "A2"
        tbuf.seek(0)
        st.download_button(
            "빈 입력 양식 다운로드",
            data=tbuf.read(),
            file_name="campaign_store_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_campaign_template",
        )
        if "campaign_upload_bytes" not in st.session_state:
            st.session_state["campaign_upload_bytes"] = b""
        if "campaign_upload_name" not in st.session_state:
            st.session_state["campaign_upload_name"] = ""
        if "campaign_run_requested" not in st.session_state:
            st.session_state["campaign_run_requested"] = False
        up_campaign = st.file_uploader(
            "작성한 템플릿 업로드 (.xlsx)",
            type=["xlsx"],
            key="campaign_upload_xlsx",
            help=f".xlsx만 가능 · 파일당 최대 {_MAX_XLSX_UPLOAD_BYTES // (1024 * 1024)}MB · 업로드 후 「작업 시작」 필요",
        )
        c1, c2 = st.columns([1.0, 1.2])
        with c1:
            if st.button("업로드 결과 초기화", key="campaign_reset_upload"):
                st.session_state["campaign_upload_bytes"] = b""
                st.session_state["campaign_upload_name"] = ""
                st.session_state["campaign_run_requested"] = False
                st.session_state.pop("campaign_upload_xlsx", None)
                st.rerun()
        with c2:
            if st.session_state.get("campaign_upload_name"):
                st.caption(f"현재 유지 중인 업로드 파일: {st.session_state['campaign_upload_name']}")

        if up_campaign is not None:
            _cerr = _xlsx_upload_size_error(up_campaign)
            if _cerr:
                st.error(_cerr)
            else:
                new_bytes = up_campaign.getvalue()
                new_name = getattr(up_campaign, "name", "uploaded.xlsx")
                prev_bytes = st.session_state.get("campaign_upload_bytes") or b""
                prev_name = st.session_state.get("campaign_upload_name") or ""
                if (new_name != prev_name) or (new_bytes != prev_bytes):
                    st.session_state["campaign_run_requested"] = False
                st.session_state["campaign_upload_bytes"] = new_bytes
                st.session_state["campaign_upload_name"] = new_name

        cached_bytes = st.session_state.get("campaign_upload_bytes") or b""
        if not cached_bytes:
            st.info(
                "빈 양식을 내려받아 «매장명»만 적은 뒤 업로드하면 됩니다. 권역은 매장 데이터에 있는 값으로 붙고, "
                f"같은 권역 안에서는 위에서부터 1·2·3… 순서로 반영됩니다(권역당 최대 {stores_per_team}곳). 빈 행은 무시됩니다.\n\n"
                "업로드만으로는 표가 만들어지지 않습니다. 아래 「작업 시작」을 눌러 주세요. "
                "페이지 새로고침·재접속 후에는 업로드와 「작업 시작」을 다시 할 수 있습니다."
            )
        else:
            run_col1, run_col2 = st.columns([1.0, 2.0])
            with run_col1:
                if st.button("작업 시작", key="campaign_run_start", type="primary", use_container_width=True):
                    st.session_state["campaign_run_requested"] = True
                    st.rerun()
            with run_col2:
                if not st.session_state.get("campaign_run_requested", False):
                    st.info(
                        "업로드가 반영되었습니다. 「작업 시작」을 누르면 아래에 결과가 생성됩니다. "
                        "(새로고침 후에는 다시 「작업 시작」이 필요할 수 있습니다.)"
                    )
            if not st.session_state.get("campaign_run_requested", False):
                st.stop()
            try:
                plan_df = parse_campaign_submission_xlsx(cached_bytes)
            except Exception as e:
                st.error(f"업로드 파일을 읽을 수 없습니다: {e}")
                st.stop()
            selected39, warns, errs = resolve_campaign_stores(
                plan_df, stores_base, expected_per_region=stores_per_team
            )
            for wmsg in warns[:6]:
                st.warning(wmsg)
            if errs:
                st.error("매장 확인이 필요한 항목이 있습니다. 아래 30건까지만 표시합니다.")
                render_table(pd.DataFrame({"확인필요": errs[:30]}), use_container_width=True, height=220)
            if selected39.empty:
                st.info("유효한 매장을 찾지 못했습니다. 매장명을 데이터와 동일하게 확인해 주세요.")
                st.stop()
            if stores_view.empty:
                st.warning(
                    "좌측 필터로 표시할 매장이 없습니다. 운영팀·권역·매장 특성을 확인해 주세요."
                )
                st.stop()
            _n_after_resolve = int(len(selected39))
            _k_sep = "\x1e"
            _sv_keys = set(
                (
                    stores_view["name"].astype(str).str.strip()
                    + _k_sep
                    + stores_view["address"].astype(str).str.strip()
                ).drop_duplicates().tolist()
            )
            selected39 = selected39.assign(
                _k=selected39["name"].astype(str).str.strip()
                + _k_sep
                + selected39["address"].astype(str).str.strip()
            )
            selected39 = (
                selected39.loc[selected39["_k"].isin(_sv_keys)]
                .drop(columns=["_k"], errors="ignore")
                .reset_index(drop=True)
            )
            if len(selected39) < _n_after_resolve:
                st.info(
                    f"좌측 «운영팀·권역»·«매장 특성 필터»에 따라 업로드 매장 중 "
                    f"{_n_after_resolve - len(selected39):,}곳을 제외했습니다. (반영 {len(selected39):,}곳)"
                )
            if selected39.empty:
                st.warning(
                    "업로드한 매장이 현재 좌측에서 선택된 매장에 없습니다. 필터를 조정하거나 업로드를 확인해 주세요."
                )
                st.stop()

            exec_df = build_campaign_execution_table(selected39, schools_use, n_near=campaign_n_near)
            dedup_df = build_school_dedup_table(exec_df)
            p1 = len(exec_df)
            p2 = dedup_df["학교명"].nunique() if not dedup_df.empty else 0
            k1, k2, k3, k4, k5 = st.columns(5)
            with k1:
                st.metric("선정 매장", f"{len(selected39):,}")
            with k2:
                st.metric("지역 수", f"{selected39['campaign_region'].nunique():,}")
            with k3:
                st.metric("지역당 평균", f"{len(selected39) / max(1, selected39['campaign_region'].nunique()):.1f}")
            with k4:
                st.metric("학교 후보(전체)", f"{p1:,}")
            with k5:
                st.metric("중복제거 학교", f"{p2:,}")
            _reg_counts = selected39.groupby("campaign_region").size()
            st.caption(
                "권역별 선정 매장 수 · "
                + " · ".join(f"{str(k)} {int(v)}곳" for k, v in _reg_counts.items())
            )

            show_cols = ["campaign_region", "campaign_rank", "address", "name"]
            show = selected39[show_cols].rename(
                columns={
                    "campaign_region": "지역",
                    "campaign_rank": "우선순위",
                    "address": "매장주소",
                    "name": "매장명",
                }
            )
            show = show[["지역", "우선순위", "매장주소", "매장명"]]
            st.markdown(f"##### 1) 취합 매장 (총 {len(selected39):,}개)")
            _b_show = BytesIO()
            with pd.ExcelWriter(_b_show, engine="openpyxl") as _w:
                _sanitize_table(show).to_excel(_w, index=False, sheet_name="취합매장")
            _b_show.seek(0)
            st.download_button(
                "엑셀 다운로드",
                data=_b_show.read(),
                file_name="summary_1_취합매장.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_summary_show",
            )
            render_table(show, use_container_width=True, height=280)
            st.markdown(
                """<div style="color:rgba(49,51,63,0.78);font-size:0.96rem;line-height:1.55;margin:0.08rem 0 0.35rem 0;">
• 정의(1) 취합 매장: 업로드한 매장명을 데이터와 맞춘 결과입니다.<br>
• 지역(권역): 매장 마스터의 권역이 붙습니다. 엑셀에 권역을 적은 경우에는 일치할 때만 반영됩니다.<br>
• 우선순위: 같은 권역 안에서는 파일 위에서부터 1, 2, 3… 순입니다(1이 가장 우선).<br>
• 이 표를 기준으로 아래 학교 통합 목록이 계산됩니다.
</div>""",
                unsafe_allow_html=True,
            )

            if exec_df.empty:
                st.info("근처 학교 계산 결과가 없습니다. 학교 필터/좌표를 확인해 주세요.")
            else:
                st.markdown("<div style='height:0.9rem;'></div>", unsafe_allow_html=True)
                st.markdown("##### 2) 학교 통합 목록(중복 방지)")
                bucket_to_regions: dict[str, set[str]] = {
                    "운영1~3": set(),
                    "운영4~6": set(),
                    "운영7~9": set(),
                    "운영10~13": set(),
                }
                for _, sr in selected39.iterrows():
                    rg = str(sr.get("campaign_region", "")).strip()
                    ot = str(sr.get("ops_team", "")).strip()
                    m = re.search(r"운영\s*(\d+)", _team_key(ot))
                    if not rg or not m:
                        continue
                    no = int(m.group(1))
                    if 1 <= no <= 3:
                        bucket_to_regions["운영1~3"].add(rg)
                    elif 4 <= no <= 6:
                        bucket_to_regions["운영4~6"].add(rg)
                    elif 7 <= no <= 9:
                        bucket_to_regions["운영7~9"].add(rg)
                    elif 10 <= no <= 13:
                        bucket_to_regions["운영10~13"].add(rg)

                # 두 필터는 동시 적용 의미가 겹치므로 하나만 선택 가능하도록 제어
                _owner_prev = str(st.session_state.get("campaign_owner_bucket_filter", "(전체)"))
                _region_prev = str(st.session_state.get("campaign_region_filter", "(전체)"))
                if _owner_prev != "(전체)" and _region_prev != "(전체)":
                    st.session_state["campaign_region_filter"] = "(전체)"
                    _region_prev = "(전체)"

                r_opts = ["(전체)"] + sorted({str(x).strip() for x in selected39["campaign_region"].tolist() if str(x).strip()})
                _flt_l, _flt_r = st.columns(2, gap="medium")
                with _flt_l:
                    owner_bucket = st.selectbox(
                        "권역 담당자 묶음 필터",
                        options=["(전체)", "운영1~3", "운영4~6", "운영7~9", "운영10~13"],
                        index=0,
                        key="campaign_owner_bucket_filter",
                        help="담당 운영팀 묶음 기준으로 해당 지역 학교만 추려볼 수 있습니다.",
                        disabled=(_region_prev != "(전체)"),
                    )
                with _flt_r:
                    r_pick = st.selectbox(
                        "지역 필터 (권역 담당자 묶음이 '(전체)'일 때 선택 가능)",
                        options=r_opts,
                        index=0,
                        key="campaign_region_filter",
                        disabled=(str(owner_bucket) != "(전체)"),
                    )
                max_store_n = max(1, int(pd.to_numeric(dedup_df["연관매장수"], errors="coerce").max()))
                _max_pr = (
                    int(selected39.groupby("campaign_region").size().max())
                    if not selected39.empty
                    else int(stores_per_team)
                )
                core_v = min(max_store_n, max(1, _max_pr))
                mid_v = min(max_store_n, max(1, core_v - 1))
                single_v = 1
                quick_options: list[str] = [f"1순위({core_v}개 이상 겹침)"]
                if mid_v not in (core_v, single_v):
                    quick_options.append(f"2순위({mid_v}개 매장 겹침)")
                quick_options.append(f"3순위({single_v}개 매장)")
                quick_options.append("직접범위")
                mode = st.segmented_control(
                    "빠른 필터",
                    options=quick_options,
                    selection_mode="single",
                    default=quick_options[0],
                    key="campaign_dedup_mode",
                    help="1·2·3순위는 연관매장수 기준으로 우선 컨택할 학교 후보를 나눈 것입니다. "
                    "경계값은 이번 업로드에서 권역별 매장 수가 가장 많은 값에 맞춰 잡힙니다.",
                )
                st.markdown(
                    f"""<div style="color:rgba(49,51,63,0.78);font-size:0.96rem;line-height:1.55;margin:0.08rem 0 0.35rem 0;">
• 1순위: 같은 학교가 여러 매장(기준 {core_v}개 이상)에서 동시에 가까운 학교로 잡힌 경우 — 우선 컨택 후보<br>
• 2순위: 같은 학교가 {mid_v}개 매장에서 겹치는 경우<br>
• 3순위: 한 매장에서만 가까운 학교로 잡힌 경우<br>
• 직접범위: 연관매장수 최소~최대를 직접 지정
</div>""",
                    unsafe_allow_html=True,
                )
                if mode == f"1순위({core_v}개 이상 겹침)":
                    store_n_range = (core_v, max_store_n)
                elif mode == f"2순위({mid_v}개 매장 겹침)":
                    store_n_range = (mid_v, mid_v)
                elif mode == f"3순위({single_v}개 매장)":
                    store_n_range = (single_v, single_v)
                else:
                    store_n_range = st.slider(
                        "연관매장수 범위 (최소~최대)",
                        min_value=1,
                        max_value=max_store_n,
                        value=(1, max_store_n),
                        step=1,
                        key="campaign_dedup_store_range",
                    )
                dedup_view = dedup_df.copy()
                _sn = pd.to_numeric(dedup_view["연관매장수"], errors="coerce").fillna(0)
                dedup_view = dedup_view[(_sn >= store_n_range[0]) & (_sn <= store_n_range[1])]
                if owner_bucket != "(전체)":
                    allowed_regions = sorted(bucket_to_regions.get(owner_bucket, set()))
                    if allowed_regions:
                        pat = "|".join(re.escape(x) for x in allowed_regions)
                        dedup_view = dedup_view[
                            dedup_view["관련지역"].astype(str).str.contains(pat, regex=True, na=False)
                        ]
                        stores_view_filtered = selected39[
                            selected39["campaign_region"].astype(str).isin(allowed_regions)
                        ].copy()
                    else:
                        dedup_view = dedup_view.iloc[0:0]
                        stores_view_filtered = selected39.iloc[0:0].copy()
                else:
                    stores_view_filtered = selected39.copy()
                if r_pick != "(전체)":
                    dedup_view = dedup_view[
                        dedup_view["관련지역"].astype(str).str.contains(re.escape(r_pick), regex=True, na=False)
                    ]
                    stores_view_filtered = stores_view_filtered[
                        stores_view_filtered["campaign_region"].astype(str) == str(r_pick)
                    ].copy()
                stores_show = (
                    stores_view_filtered[["campaign_region", "campaign_rank", "name", "address"]]
                    .rename(
                        columns={
                            "campaign_region": "지역",
                            "campaign_rank": "우선순위",
                            "name": "매장명",
                            "address": "매장주소",
                        }
                    )
                    .sort_values(["지역", "우선순위", "매장명"])
                    .reset_index(drop=True)
                )
                with st.expander("반영 매장 전체 목록 보기", expanded=False):
                    _b_stores = BytesIO()
                    with pd.ExcelWriter(_b_stores, engine="openpyxl") as _w:
                        _sanitize_table(stores_show).to_excel(_w, index=False, sheet_name="반영매장목록")
                    _b_stores.seek(0)
                    st.download_button(
                        "엑셀 다운로드",
                        data=_b_stores.read(),
                        file_name="summary_3_반영매장목록.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_summary_stores_show",
                    )
                    render_table(stores_show, use_container_width=True, height=220)
                exec_linked = exec_df.copy()
                if owner_bucket != "(전체)":
                    allowed_regions = sorted(bucket_to_regions.get(owner_bucket, set()))
                    if allowed_regions:
                        exec_linked = exec_linked[exec_linked["권역"].astype(str).isin(allowed_regions)]
                    else:
                        exec_linked = exec_linked.iloc[0:0]
                if r_pick != "(전체)":
                    exec_linked = exec_linked[exec_linked["권역"].astype(str) == str(r_pick)]
                _sn_exec = pd.to_numeric(
                    exec_linked.groupby(["학교명", "학교주소"])["매장명"].transform("nunique"),
                    errors="coerce",
                ).fillna(0)
                _sn_view = pd.to_numeric(
                    exec_linked.groupby(["학교명", "학교주소"])["매장명"].transform("nunique"),
                    errors="coerce",
                ).fillna(0)
                exec_reflected_any = exec_linked[
                    (_sn_exec >= int(core_v)) | (_sn_exec == int(mid_v)) | (_sn_exec == int(single_v))
                ]
                union_priority_store_n = (
                    int(exec_reflected_any["매장명"].astype(str).str.strip().nunique())
                    if not exec_reflected_any.empty
                    else 0
                )
                exec_for_current_view = exec_linked[(_sn_view >= store_n_range[0]) & (_sn_view <= store_n_range[1])]
                current_view_store_n = int(exec_for_current_view["매장명"].astype(str).str.strip().nunique())
                uploaded_store_n = int(selected39["name"].astype(str).str.strip().nunique())
                st.caption(
                    f"빠른필터 반영 매장수 {current_view_store_n:,} / "
                    f"선택 필터 기준(1·2·3순위 매장 합계) {union_priority_store_n:,} / "
                    f"업로드된 전체 매장수 {uploaded_store_n:,}"
                )
                st.caption(f"현재 목록 학교 수: {len(dedup_view):,}개")
                dedup_view_disp = dedup_view.copy()
                dedup_view_disp["집계기준"] = f"매장별 최근접 상위 {int(campaign_n_near)}개"
                pref_cols = [
                    "학교명",
                    "학교유형",
                    "캠퍼스구분",
                    "학교주소",
                    "연관매장수",
                    "매장목록",
                    "집계기준",
                    "최대직선거리(km)",
                    "관련지역",
                ]
                rest_cols = [c for c in dedup_view_disp.columns if c not in pref_cols]
                dedup_view_disp = dedup_view_disp[pref_cols + rest_cols]
                _b_dedup = BytesIO()
                with pd.ExcelWriter(_b_dedup, engine="openpyxl") as _w:
                    _sanitize_table(dedup_view_disp).to_excel(_w, index=False, sheet_name="학교통합목록")
                _b_dedup.seek(0)
                st.download_button(
                    "엑셀 다운로드",
                    data=_b_dedup.read(),
                    file_name="summary_2_학교통합목록.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_summary_dedup",
                )
                render_table(dedup_view_disp, use_container_width=True, height=360)
                st.markdown(
                    f"""<div style="color:rgba(49,51,63,0.78);font-size:0.96rem;line-height:1.55;margin:0.08rem 0 0.35rem 0;">
• 정의(2) 학교 통합 목록: 취합 매장 전체를 합쳐 중복 학교를 1번만 남긴 목록입니다.<br>
• 연관매장수: 해당 학교를 가까운 학교로 포함한 매장 수입니다(현재 매장당 상위 {int(campaign_n_near)}개 기준).<br>
• 최대직선거리: 연관 매장 중 해당 학교와 가장 먼 매장까지의 직선거리입니다.
</div>""",
                    unsafe_allow_html=True,
                )
                linked_store_names = set(exec_reflected_any["매장명"].astype(str).str.strip().tolist())
                store_school_counts = (
                    exec_linked.groupby("매장명")
                    .apply(lambda g: int(g[["학교명", "학교주소"]].drop_duplicates().shape[0]))
                    .to_dict()
                    if not exec_linked.empty
                    else {}
                )
                unlinked_df = stores_view_filtered[
                    ~stores_view_filtered["name"].astype(str).str.strip().isin(linked_store_names)
                ][["campaign_region", "campaign_rank", "name", "address"]].rename(
                    columns={
                        "campaign_region": "지역",
                        "campaign_rank": "우선순위",
                        "name": "매장명",
                        "address": "매장주소",
                    }
                ).sort_values(["지역", "우선순위", "매장명"]).reset_index(drop=True)
                if not unlinked_df.empty:
                    unlinked_df["매칭학교수"] = (
                        unlinked_df["매장명"].map(lambda x: int(store_school_counts.get(str(x).strip(), 0)))
                    )
                    unlinked_df["사유"] = np.where(
                        unlinked_df["매칭학교수"] <= 0,
                        "현재 필터 기준 매칭 학교 없음",
                        "1·2·3순위 기준(연관매장수) 밖",
                    )
                st.markdown("##### 3) 미반영 매장 점검")
                if unlinked_df.empty:
                    st.success("현재 지역/담당자 조건에서 모든 매장이 1·2·3순위 중 하나에는 반영되었습니다.")
                else:
                    st.warning(f"1·2·3순위 어디에도 반영되지 않은 매장 {len(unlinked_df):,}개")
                    _b_un = BytesIO()
                    with pd.ExcelWriter(_b_un, engine="openpyxl") as _w:
                        _sanitize_table(unlinked_df).to_excel(_w, index=False, sheet_name="미반영매장")
                    _b_un.seek(0)
                    st.download_button(
                        "엑셀 다운로드",
                        data=_b_un.read(),
                        file_name="summary_3_미반영매장.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_summary_unlinked",
                    )
                    render_table(unlinked_df, use_container_width=True, height=220)
            b_exec = BytesIO()
            with pd.ExcelWriter(b_exec, engine="openpyxl") as w:
                _sanitize_table(show).to_excel(w, index=False, sheet_name="취합매장")
                if not exec_df.empty:
                    _sanitize_table(dedup_df).to_excel(w, index=False, sheet_name="학교중복통합")
            b_exec.seek(0)
            st.download_button(
                "업로드 결과 엑셀 다운로드",
                data=b_exec.read(),
                file_name="sanhak_summary_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
